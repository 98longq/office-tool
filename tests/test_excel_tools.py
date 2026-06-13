import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from office_tool.excel import clean_workbook, inspect_workbook


class ExcelToolTests(unittest.TestCase):
    def test_inspect_workbook_counts_cells_and_sheets(self):
        with tempfile.TemporaryDirectory(prefix="office_tool_excel_") as tmp:
            path = Path(tmp) / "book.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "数据"
            sheet["A1"] = "姓名"
            sheet["B2"] = "张三"
            workbook.create_sheet("空表")
            workbook.save(path)

            summary = inspect_workbook(path)

        self.assertEqual(len(summary.sheets), 2)
        self.assertEqual(summary.sheets[0].name, "数据")
        self.assertEqual(summary.sheets[0].non_empty_cells, 2)

    def test_clean_workbook_trims_text_and_removes_empty_rows(self):
        with tempfile.TemporaryDirectory(prefix="office_tool_excel_clean_") as tmp:
            source = Path(tmp) / "raw.xlsx"
            output = Path(tmp) / "clean.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "  文本  "
            sheet["A3"] = "尾行"
            workbook.save(source)

            summary = clean_workbook(source, output)
            cleaned = load_workbook(output)
            sheet = cleaned.active

            self.assertEqual(sheet["A1"].value, "文本")
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(summary.sheets[0].rows, 2)
            cleaned.close()


if __name__ == "__main__":
    unittest.main()
