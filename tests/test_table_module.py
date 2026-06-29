import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from office_tool.table_audit import TableWorkbookInspector, collect_table_inputs, detect_header_row
from office_tool.table_merge import merge_by_columns, merge_same_layout
from office_tool.table_models import SourceColumnMapping, TableMergeOptions


class TableModuleTests(unittest.TestCase):
    def test_inspector_detects_header_below_merged_title(self):
        with tempfile.TemporaryDirectory(prefix="office_tool_table_inspect_") as tmp:
            path = Path(tmp) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "填报表"
            sheet.merge_cells("A1:C1")
            sheet["A1"] = "部门任务填报表"
            sheet.append(["任务", "办理情况", "备注"])
            sheet.append(["任务A", "", ""])
            workbook.save(path)

            report = TableWorkbookInspector().inspect_path(path)

        self.assertEqual(report.sheets[0].header_row, 2)
        self.assertEqual(report.sheets[0].headers[:3], ["任务", "办理情况", "备注"])
        self.assertEqual(report.sheets[0].merged_ranges, ["A1:C1"])

    def test_merge_by_columns_appends_distinct_source_values(self):
        with tempfile.TemporaryDirectory(prefix="office_tool_table_merge_") as tmp:
            root = Path(tmp)
            master = root / "master.xlsx"
            source = root / "source.xlsx"
            output = root / "merged.xlsx"

            master_wb = Workbook()
            master_ws = master_wb.active
            master_ws.title = "总表"
            master_ws.merge_cells("A1:C1")
            master_ws["A1"] = "任务汇总"
            master_ws.append(["任务", "办理情况", "备注"])
            master_ws.append(["任务A", "已收到", ""])
            master_ws.append(["任务B", "", ""])
            master_wb.save(master)

            source_wb = Workbook()
            source_ws = source_wb.active
            source_ws.title = "办公室"
            source_ws.append(["任务", "回复"])
            source_ws.append(["任务A", "已收到"])
            source_ws.append(["任务A", "正在推进"])
            source_ws.append(["任务B", "已办结"])
            source_ws.append(["任务C", "无需处理"])
            source_wb.save(source)

            report = merge_by_columns(
                TableMergeOptions(
                    master_path=master,
                    output_path=output,
                    master_sheet="总表",
                    master_key_column="任务",
                    master_target_column="办理情况",
                    sources=[
                        SourceColumnMapping(
                            path=source,
                            sheet="办公室",
                            key_column="任务",
                            value_column="回复",
                        )
                    ],
                )
            )

            merged = load_workbook(output)
            sheet = merged["总表"]

        self.assertEqual(sheet["B3"].value, "已收到\n正在推进")
        self.assertEqual(sheet["B4"].value, "已办结")
        self.assertEqual(report.stats["updated_cells"], 2)
        self.assertIn("duplicate_source_key", {finding.code for finding in report.findings})
        self.assertIn("unused_source_value", {finding.code for finding in report.findings})

    def test_merge_by_columns_accepts_different_source_headers(self):
        with tempfile.TemporaryDirectory(prefix="office_tool_table_special_headers_") as tmp:
            root = Path(tmp)
            master = root / "master.xlsx"
            source_a = root / "source_a.xlsx"
            source_b = root / "source_b.xlsx"
            output = root / "merged.xlsx"

            master_wb = Workbook()
            master_ws = master_wb.active
            master_ws.title = "总表"
            master_ws.append(["任务详情", "任务回复"])
            master_ws.append(["任务一", ""])
            master_ws.append(["任务二", ""])
            master_wb.save(master)

            source_a_wb = Workbook()
            source_a_ws = source_a_wb.active
            source_a_ws.title = "部门A"
            source_a_ws.append(["任务", "回复"])
            source_a_ws.append(["任务一", "A 已落实"])
            source_a_wb.save(source_a)

            source_b_wb = Workbook()
            source_b_ws = source_b_wb.active
            source_b_ws.title = "部门B"
            source_b_ws.append(["下达任务", "落实"])
            source_b_ws.append(["任务一", "B 已落实"])
            source_b_ws.append(["任务二", "B 已反馈"])
            source_b_wb.save(source_b)

            report = merge_by_columns(
                TableMergeOptions(
                    master_path=master,
                    output_path=output,
                    master_sheet="总表",
                    master_key_column="任务详情",
                    master_target_column="任务回复",
                    sources=[
                        SourceColumnMapping(path=source_a, sheet="部门A", key_column="任务", value_column="回复"),
                        SourceColumnMapping(path=source_b, sheet="部门B", key_column="下达任务", value_column="落实"),
                    ],
                )
            )

            merged = load_workbook(output)
            sheet = merged["总表"]

        self.assertEqual(sheet["B2"].value, "A 已落实\nB 已落实")
        self.assertEqual(sheet["B3"].value, "B 已反馈")
        self.assertEqual(report.stats["updated_cells"], 2)

    def test_merge_same_layout_appends_by_cell_position(self):
        with tempfile.TemporaryDirectory(prefix="office_tool_table_same_layout_") as tmp:
            root = Path(tmp)
            master = root / "master.xlsx"
            source_a = root / "source_a.xlsx"
            source_b = root / "source_b.xlsx"
            output = root / "merged.xlsx"

            master_wb = Workbook()
            master_ws = master_wb.active
            master_ws.title = "总表"
            master_ws.append(["任务", "办公室", "财务部"])
            master_ws.append(["任务一", "", ""])
            master_ws.append(["任务二", "", ""])
            master_wb.save(master)

            source_a_wb = Workbook()
            source_a_ws = source_a_wb.active
            source_a_ws.append(["任务", "办公室", "财务部"])
            source_a_ws.append(["任务一", "已完成", ""])
            source_a_ws.append(["任务二", "推进中", ""])
            source_a_wb.save(source_a)

            source_b_wb = Workbook()
            source_b_ws = source_b_wb.active
            source_b_ws.append(["任务", "办公室", "财务部"])
            source_b_ws.append(["任务一", "", "已反馈"])
            source_b_ws.append(["任务二", "", "推进中"])
            source_b_wb.save(source_b)

            report = merge_same_layout(master, [source_a, source_b], output)
            merged = load_workbook(output)
            sheet = merged["总表"]

        self.assertEqual(sheet["B2"].value, "已完成")
        self.assertEqual(sheet["C2"].value, "已反馈")
        self.assertEqual(sheet["B3"].value, "推进中")
        self.assertEqual(sheet["C3"].value, "推进中")
        self.assertEqual(report.stats["updated_cells"], 4)

    def test_collect_table_inputs_skips_temporary_files(self):
        with tempfile.TemporaryDirectory(prefix="office_tool_table_collect_") as tmp:
            root = Path(tmp)
            (root / "a.xlsx").write_bytes(b"placeholder")
            (root / "~$a.xlsx").write_bytes(b"placeholder")
            (root / "skip.docx").write_bytes(b"placeholder")

            paths = collect_table_inputs([root])

        self.assertEqual([path.name for path in paths], ["a.xlsx"])


if __name__ == "__main__":
    unittest.main()
