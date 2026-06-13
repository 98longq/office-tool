"""Basic Excel helper tools."""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class SheetSummary:
    name: str
    rows: int
    columns: int
    non_empty_cells: int
    merged_ranges: int


@dataclass
class WorkbookSummary:
    path: str
    sheets: list[SheetSummary] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {"path": self.path, "sheets": [asdict(sheet) for sheet in self.sheets]}


def inspect_workbook(path: str | Path) -> WorkbookSummary:
    source = Path(path).expanduser().resolve()
    workbook = load_workbook(source, read_only=False, data_only=False)
    summary = WorkbookSummary(path=str(source))
    for sheet in workbook.worksheets:
        non_empty = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    non_empty += 1
        summary.sheets.append(
            SheetSummary(
                name=sheet.title,
                rows=sheet.max_row,
                columns=sheet.max_column,
                non_empty_cells=non_empty,
                merged_ranges=len(sheet.merged_cells.ranges),
            )
        )
    workbook.close()
    return summary


def clean_workbook(
    input_path: str | Path,
    output_path: str | Path,
    trim_text: bool = True,
    remove_empty_rows: bool = True,
) -> WorkbookSummary:
    source = Path(input_path).expanduser().resolve()
    output = Path(output_path).expanduser().resolve()
    workbook = load_workbook(source)
    for sheet in workbook.worksheets:
        if trim_text:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = cell.value.strip()
        if remove_empty_rows:
            for row_idx in range(sheet.max_row, 0, -1):
                if all(sheet.cell(row_idx, col).value in (None, "") for col in range(1, sheet.max_column + 1)):
                    sheet.delete_rows(row_idx, 1)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    return inspect_workbook(output)
