"""Tkinter desktop GUI for OfficeTool."""

from __future__ import annotations

import json
import os
import struct
import sys
import tempfile
import threading
from copy import deepcopy
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from .ai import DeepSeekTextReviewer
from .config import AIReviewOptions, OfficeToolConfig
from .services import audit_many, format_many, summarize_results
from .table_audit import TableWorkbookInspector, _load_workbook, collect_table_inputs, inspect_workbook, read_sheet_preview
from .table_merge import merge_by_columns, merge_same_layout
from .table_models import SourceColumnMapping, TableMergeOptions
from .profile_store import ConfigProfileStore
from .secret_store import protect_secret, unprotect_secret


COLOR = {
    "bg": "#f5f8fc",
    "surface": "#ffffff",
    "surface_alt": "#edf4ff",
    "surface_tint": "#dceaff",
    "border": "#c9d8ec",
    "border_strong": "#8fb0da",
    "text": "#12233a",
    "muted": "#5f6f85",
    "subtle": "#8795a8",
    "primary": "#2563eb",
    "primary_hover": "#1d4ed8",
    "primary_soft": "#dbeafe",
    "accent": "#0f4c81",
    "danger": "#b93a32",
    "warning": "#a66a12",
    "success": "#147a57",
    "white": "#ffffff",
}

FONT_FAMILY = "Microsoft YaHei UI"
MONO_FAMILY = "Cascadia Mono"
FIELD_MIN_WIDTH = 320
AI_PROFILE_DIR = Path.home() / ".office_tool"
AI_PROFILE_FILE = AI_PROFILE_DIR / "ai_profiles.json"
CONFIG_PROFILE_DIR = Path(os.environ.get("APPDATA", str(Path.home() / ".office_tool"))) / "OfficeTool" / "profiles"
BUILTIN_SCHEMES = ("普通公文", "红头文件", "红头文件（函）", "红头文件（会议纪要）")
FONT_SIZE_LABELS = {
    "初号": 42,
    "小初": 36,
    "一号": 26,
    "小一": 24,
    "二号": 22,
    "小二": 18,
    "三号": 16,
    "小三": 15,
    "四号": 14,
    "小四": 12,
    "五号": 10.5,
    "小五": 9,
}


def _font_size_label(size_pt: float) -> str:
    for label, value in FONT_SIZE_LABELS.items():
        if abs(float(size_pt) - value) < 0.01:
            return label
    return f"{float(size_pt):g}"


def _font_size_value(raw: str) -> float:
    value = raw.strip()
    if value in FONT_SIZE_LABELS:
        return float(FONT_SIZE_LABELS[value])
    return float(value)


def _attached_popup_geometry(
    *,
    anchor_x: int,
    anchor_y: int,
    anchor_width: int,
    anchor_height: int,
    requested_height: int,
    screen_x: int,
    screen_y: int,
    screen_width: int,
    screen_height: int,
) -> tuple[int, int, int, int]:
    """Return an on-screen popup rectangle attached to the anchor widget."""
    available_width = max(1, screen_width)
    width = min(max(1, anchor_width), available_width)
    height = max(1, min(requested_height, max(1, screen_height)))

    screen_right = screen_x + available_width
    screen_bottom = screen_y + max(1, screen_height)
    x = anchor_x + anchor_width - width
    x = max(screen_x, min(x, screen_right - width))

    below_y = anchor_y + anchor_height + 2
    above_y = anchor_y - height - 2
    y = above_y if below_y + height > screen_bottom and above_y >= screen_y else below_y
    y = max(screen_y, min(y, screen_bottom - height))
    return width, height, x, y


def _widget_is_within(widget: tk.Misc | None, ancestor: tk.Misc | None) -> bool:
    while widget is not None:
        if widget is ancestor:
            return True
        widget = getattr(widget, "master", None)
    return False


def _attach_tooltip(widget: tk.Widget, text: str) -> None:
    popup: tk.Toplevel | None = None

    def show(_event=None) -> None:
        nonlocal popup
        if popup is not None:
            return
        popup = tk.Toplevel(widget)
        popup.wm_overrideredirect(True)
        popup.wm_geometry(f"+{widget.winfo_rootx() + 8}+{widget.winfo_rooty() + widget.winfo_height() + 5}")
        tk.Label(
            popup,
            text=text,
            bg=COLOR["text"],
            fg=COLOR["white"],
            padx=8,
            pady=4,
            font=(FONT_FAMILY, 9),
        ).pack()

    def hide(_event=None) -> None:
        nonlocal popup
        if popup is not None:
            popup.destroy()
            popup = None

    widget.bind("<Enter>", show, add="+")
    widget.bind("<Leave>", hide, add="+")


def _install_text_placeholder(widget: tk.Text, placeholder: str) -> None:
    widget._office_tool_placeholder = placeholder
    widget._office_tool_placeholder_active = False

    def show() -> None:
        if widget.get("1.0", tk.END).strip():
            return
        widget.configure(fg=COLOR["subtle"])
        widget.insert("1.0", placeholder)
        widget.edit_reset()
        widget._office_tool_placeholder_active = True

    def focus_in(_event=None) -> None:
        if not widget._office_tool_placeholder_active:
            return
        widget.delete("1.0", tk.END)
        widget.edit_reset()
        widget.configure(fg=COLOR["text"])
        widget._office_tool_placeholder_active = False

    def focus_out(_event=None) -> None:
        if not widget.get("1.0", tk.END).strip():
            show()

    widget.bind("<FocusIn>", focus_in, add="+")
    widget.bind("<FocusOut>", focus_out, add="+")
    show()


def _text_without_placeholder(widget: tk.Text | None) -> str:
    if widget is None or getattr(widget, "_office_tool_placeholder_active", False):
        return ""
    return widget.get("1.0", tk.END).strip()


def _first_sheet_name(path: str | Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("表格功能需要安装 openpyxl。") from exc
    workbook = load_workbook(path, read_only=True)
    return workbook.sheetnames[0]

FINDING_TYPE_BY_CODE = {
    "missing_title": "结构缺失",
    "missing_document_number": "结构缺失",
    "missing_signer": "结构缺失",
    "missing_main_send": "结构缺失",
    "missing_date": "结构缺失",
    "document_number_after_title": "格式错误",
    "attachment_note_after_date": "格式错误",
    "possible_unrecognized_red_head": "格式提示",
}

AI_TYPE_KEYWORDS = {
    "格式": "格式错误",
    "版式": "格式错误",
    "漏字": "漏字错字",
    "错字": "漏字错字",
    "别字": "漏字错字",
    "语病": "语句不当",
    "不通顺": "语句不当",
    "措辞": "语句不当",
    "表述": "语句不当",
    "歧义": "语义歧义",
    "一致": "前后不一致",
    "矛盾": "前后不一致",
    "日期": "前后不一致",
    "数字": "前后不一致",
    "风险": "风险表述",
    "承诺": "风险表述",
}


def _profile_label(profile: str) -> str:
    return {
        "auto": "普通公文",
        "standard": "普通公文",
        "red_head": "红头文件",
        "letter_head": "红头文件（函）",
        "meeting_minutes": "红头文件（会议纪要）",
    }.get(profile, "普通公文")


def _default_gui_config(ai_options: AIReviewOptions | None = None) -> OfficeToolConfig:
    config = OfficeToolConfig()
    config.audit.profile = "standard"
    if ai_options is not None:
        config.ai_review = deepcopy(ai_options)
    return config

def _enable_dpi_awareness() -> None:
    """Make ttk controls respect Windows DPI scaling."""
    if sys.platform != "win32":
        return
    try:
        from ctypes import windll

        try:
            windll.shcore.SetProcessDpiAwareness(2)
            return
        except (AttributeError, OSError):
            pass
        try:
            windll.shcore.SetProcessDpiAwareness(1)
            return
        except (AttributeError, OSError):
            pass
        try:
            windll.user32.SetProcessDPIAware()
        except (AttributeError, OSError):
            pass
    except ImportError:
        pass


def _make_app_icon_photo(size: int = 32) -> tk.PhotoImage:
    photo = tk.PhotoImage(width=size, height=size)
    photo.put(COLOR["surface_alt"], to=(0, 0, size, size))
    photo.put(COLOR["primary"], to=(5, 5, size - 5, size - 5))
    photo.put(COLOR["accent"], to=(10, 10, size - 10, size - 10))
    photo.put(COLOR["white"], to=(14, 14, size - 14, size - 14))
    return photo


def _write_app_icon(path: str, size: int = 32) -> None:
    layers = (
        ((0, 0, size, size), (245, 248, 252)),
        ((5, 5, size - 5, size - 5), (37, 99, 235)),
        ((10, 10, size - 10, size - 10), (15, 76, 129)),
        ((14, 14, size - 14, size - 14), (255, 255, 255)),
    )
    header = struct.pack("<HHH", 0, 1, 1)
    image_size = 40 + size * size * 4 + size * size // 8
    entry = struct.pack("<BBBBHHII", size & 0xFF, size & 0xFF, 0, 0, 1, 32, image_size, 22)
    info = struct.pack("<IiiHHIIiiII", 40, size, size * 2, 1, 32, 0, 0, 0, 0, 0, 0)
    pixels = bytearray()
    for y in range(size - 1, -1, -1):
        for x in range(size):
            color = layers[0][1]
            for (x0, y0, x1, y1), rgb in layers:
                if x0 <= x < x1 and y0 <= y < y1:
                    color = rgb
                    break
            b, g, r = color
            pixels.extend([b & 0xFF, g & 0xFF, r & 0xFF, 0xFF])
    with open(path, "wb") as f:
        f.write(header)
        f.write(entry)
        f.write(info)
        f.write(bytes(pixels))
        f.write(bytes(size * size // 8))


def _install_app_icon(root: tk.Tk) -> tk.PhotoImage:
    photo = _make_app_icon_photo()
    root.iconphoto(True, photo)
    try:
        icon_path = os.path.join(tempfile.gettempdir(), "office_tool_app_icon.ico")
        _write_app_icon(icon_path)
        root.iconbitmap(default=icon_path)
    except Exception:
        pass
    return photo


def _make_scrollable(parent: tk.Misc) -> tuple[ttk.Frame, ttk.Frame]:
    wrapper = ttk.Frame(parent, style="Card.TFrame")
    wrapper.columnconfigure(0, weight=1)
    wrapper.rowconfigure(0, weight=1)

    canvas = tk.Canvas(wrapper, highlightthickness=0, borderwidth=0, bg=COLOR["surface"])
    canvas.grid(row=0, column=0, sticky="nsew")
    scrollbar = ttk.Scrollbar(wrapper, orient="vertical", command=canvas.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")
    canvas.configure(yscrollcommand=scrollbar.set)

    inner = ttk.Frame(canvas, style="Card.TFrame", padding=(24, 20, 24, 20))
    window_id = canvas.create_window((0, 0), window=inner, anchor="nw")

    def _resize(event) -> None:
        canvas.itemconfig(window_id, width=event.width)

    def _update_region(_event) -> None:
        canvas.configure(scrollregion=canvas.bbox("all"))

    def _on_wheel(event) -> None:
        delta = getattr(event, "delta", 0)
        if delta:
            canvas.yview_scroll(int(-1 * (delta / 120)), "units")
            return
        number = getattr(event, "num", 0)
        if number == 4:
            canvas.yview_scroll(-1, "units")
        elif number == 5:
            canvas.yview_scroll(1, "units")

    canvas.bind("<Configure>", _resize)
    inner.bind("<Configure>", _update_region)
    canvas.bind("<Enter>", lambda _e: canvas.bind_all("<MouseWheel>", _on_wheel))
    canvas.bind("<Leave>", lambda _e: canvas.unbind_all("<MouseWheel>"))
    canvas.bind("<Button-4>", _on_wheel)
    canvas.bind("<Button-5>", _on_wheel)
    inner.columnconfigure(0, weight=1, minsize=FIELD_MIN_WIDTH)
    return wrapper, inner


class OfficeToolGUI:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("办公助手")
        self.root.geometry("1480x900")
        self.root.minsize(1180, 760)
        self.config = _default_gui_config()
        self._app_icon = _install_app_icon(self.root)

        self.document_paths: list[Path] = []
        self.table_paths: list[Path] = []
        self.table_master_path: Path | None = None
        self.table_selected_source_path: Path | None = None
        self.table_workbook_infos: dict[Path, list] = {}
        self.table_source_overrides: dict[Path, dict[str, str]] = {}
        self.direct_text: tk.Text | None = None
        self.doc_list_placeholder: tk.Label | None = None
        self.table_list_placeholder: tk.Label | None = None
        self.table_tree_placeholder: tk.Label | None = None
        self.result_tree_placeholder: tk.Label | None = None
        self.doc_output_dir = tk.StringVar()
        self.doc_report_dir = tk.StringVar()
        self.table_master_file = tk.StringVar()
        self.table_master_sheet = tk.StringVar(value="总表")
        self.table_master_key_column = tk.StringVar(value="任务名称")
        self.table_master_target_column = tk.StringVar(value="办理情况")
        self.table_source_sheet = tk.StringVar()
        self.table_source_key_column = tk.StringVar(value="任务名称")
        self.table_source_value_column = tk.StringVar(value="回复内容")
        self.table_selected_source_sheet = tk.StringVar()
        self.table_selected_source_key_column = tk.StringVar()
        self.table_selected_source_value_column = tk.StringVar()
        self.table_fuzzy_enabled = tk.BooleanVar(value=False)
        self.table_fuzzy_threshold = tk.IntVar(value=90)
        self.table_output_file = tk.StringVar()
        self.status_text = tk.StringVar(value="就绪")
        self.ai_enabled = tk.BooleanVar(value=False)
        self.ai_base_url = tk.StringVar()
        self.ai_model = tk.StringVar(value="deepseek-chat")
        self.ai_api_key = tk.StringVar()
        self.ai_key_env = tk.StringVar(value="DEEPSEEK_API_KEY")
        self.ai_auth_prefix = tk.StringVar(value="Bearer")
        self.ai_stream = tk.BooleanVar(value=False)
        self.ai_toggle_text = tk.StringVar(value="启用 AI")
        self.ai_profile_name = tk.StringVar(value="默认配置")
        self.ai_profiles: dict[str, AIReviewOptions] = {}
        self.page_vars: dict[str, tk.StringVar] = {}
        self.audit_vars: dict[str, tk.BooleanVar] = {}
        self.format_vars: dict[str, tk.BooleanVar] = {}
        self.style_vars: dict[str, tuple[tk.StringVar, tk.StringVar]] = {}
        self.generation_vars: dict[str, tk.Variable] = {}
        self.scheme_var = tk.StringVar(value="普通公文")
        self.config_profile_store = ConfigProfileStore(CONFIG_PROFILE_DIR)
        self.config_profiles = self.config_profile_store.load_all()
        self.scheme_box: ttk.Combobox | None = None
        self.config_profile_button: ttk.Button | None = None
        self.config_profile_popup: tk.Toplevel | None = None
        self.config_profile_anchor: tk.Widget | None = None
        self.config_popup_bind_id: str | None = None
        self.config_popup_click_bind_ids: dict[str, str] = {}
        self._refresh_generation_panel = lambda: None
        self._apply_scheme_callback = lambda _name: None
        self.result_details: dict[str, str] = {}
        self.table_details: dict[str, str] = {}
        self.table_master_sheet_box: ttk.Combobox | None = None
        self.table_master_key_box: ttk.Combobox | None = None
        self.table_master_target_box: ttk.Combobox | None = None
        self.table_source_sheet_box: ttk.Combobox | None = None
        self.table_source_key_box: ttk.Combobox | None = None
        self.table_source_value_box: ttk.Combobox | None = None
        self.table_selected_source_sheet_box: ttk.Combobox | None = None
        self.table_selected_source_key_box: ttk.Combobox | None = None
        self.table_selected_source_value_box: ttk.Combobox | None = None
        self.table_rules_tree: ttk.Treeview | None = None
        self.table_preview_canvas: tk.Canvas | None = None
        self.table_preview_frame: ttk.Frame | None = None
        self.table_preview_placeholder: tk.Label | None = None
        self.table_advanced_visible = tk.BooleanVar(value=False)
        self.table_advanced_frame: ttk.Frame | None = None
        self.table_advanced_button_text = tk.StringVar(value="展开高级列匹配")

        self._init_config_vars()
        self._load_ai_profiles()
        self._configure_theme()
        self._build()

    def _configure_theme(self) -> None:
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        c = COLOR
        self.root.configure(bg=c["bg"])
        style.configure(".", font=(FONT_FAMILY, 10), background=c["bg"], foreground=c["text"])
        style.configure("TFrame", background=c["bg"])
        style.configure("Card.TFrame", background=c["surface"], relief="flat")
        style.configure("Tint.TFrame", background=c["surface_alt"])
        style.configure("Toolbar.TFrame", background=c["surface"])
        style.configure("TLabel", background=c["bg"], foreground=c["text"])
        style.configure("Card.TLabel", background=c["surface"], foreground=c["text"])
        style.configure("Muted.TLabel", background=c["surface"], foreground=c["muted"], font=(FONT_FAMILY, 9))
        style.configure("Tint.TLabel", background=c["surface_alt"], foreground=c["accent"], font=(FONT_FAMILY, 9, "bold"))
        style.configure("TintMuted.TLabel", background=c["surface_alt"], foreground=c["muted"], font=(FONT_FAMILY, 9))
        style.configure("PageSub.TLabel", background=c["bg"], foreground=c["muted"], font=(FONT_FAMILY, 10))
        style.configure("CardTitle.TLabel", background=c["surface"], foreground=c["text"], font=(FONT_FAMILY, 13, "bold"))
        style.configure("Section.TLabel", background=c["surface"], foreground=c["text"], font=(FONT_FAMILY, 11, "bold"))
        style.configure("Status.TLabel", background=c["surface_alt"], foreground=c["muted"], font=(FONT_FAMILY, 9))

        style.configure("TButton", padding=(12, 7), background=c["surface"], foreground=c["text"], bordercolor=c["border"], lightcolor=c["border"], darkcolor=c["border"], relief="flat")
        style.map("TButton", background=[("active", c["surface_alt"]), ("pressed", c["surface_tint"])], foreground=[("disabled", c["subtle"])])
        style.configure("Primary.TButton", padding=(16, 9), background=c["primary"], foreground=c["white"], bordercolor=c["primary"], relief="flat", font=(FONT_FAMILY, 10, "bold"))
        style.map("Primary.TButton", background=[("active", c["primary_hover"]), ("pressed", c["primary_hover"])], foreground=[("active", c["white"]), ("pressed", c["white"])])
        style.configure("Secondary.TButton", padding=(16, 9), background=c["surface_alt"], foreground=c["accent"], bordercolor=c["border"], relief="flat", font=(FONT_FAMILY, 10))
        style.map("Secondary.TButton", background=[("active", c["surface_tint"]), ("pressed", c["surface_tint"])], foreground=[("active", c["accent"]), ("pressed", c["accent"])])
        style.configure("Subtle.TButton", padding=(16, 9), background=c["surface"], foreground=c["muted"], bordercolor=c["border"], relief="flat", font=(FONT_FAMILY, 10))
        style.map("Subtle.TButton", background=[("active", c["surface_alt"]), ("pressed", c["surface_tint"])], foreground=[("active", c["accent"]), ("pressed", c["accent"])])
        style.configure("Workbench.TButton", padding=(16, 10), background=c["surface_alt"], foreground=c["accent"], bordercolor=c["border"], relief="flat", font=(FONT_FAMILY, 10, "bold"))
        style.map("Workbench.TButton", background=[("active", c["surface_tint"]), ("pressed", c["primary_soft"])], foreground=[("active", c["primary"]), ("pressed", c["primary"])])
        style.configure("Quiet.TButton", padding=(10, 6), background=c["surface"], foreground=c["muted"], bordercolor=c["border"], relief="flat")
        style.map("Quiet.TButton", background=[("active", c["surface_alt"]), ("pressed", c["surface_tint"])], foreground=[("active", c["text"])])
        style.configure("Config.TMenubutton", padding=(14, 9), background=c["surface_alt"], foreground=c["accent"], bordercolor=c["border_strong"], relief="flat", font=(FONT_FAMILY, 10, "bold"))
        style.map("Config.TMenubutton", background=[("active", c["surface_tint"]), ("pressed", c["primary_soft"])], foreground=[("active", c["primary"])], bordercolor=[("active", c["primary"])])
        style.configure("Config.TButton", padding=(14, 9), background=c["surface_alt"], foreground=c["accent"], bordercolor=c["border_strong"], relief="flat", font=(FONT_FAMILY, 10, "bold"))
        style.map("Config.TButton", background=[("active", c["surface_tint"]), ("pressed", c["primary_soft"])], foreground=[("active", c["primary"])], bordercolor=[("active", c["primary"])])
        style.configure("Dropdown.TButton", padding=(10, 6), background=c["surface_alt"], foreground=c["accent"], bordercolor=c["border"], relief="flat", font=(FONT_FAMILY, 9))
        style.map("Dropdown.TButton", background=[("active", c["surface_tint"]), ("pressed", c["primary_soft"])], foreground=[("active", c["primary"])])
        style.configure("ResultAction.TButton", padding=(11, 6), background=c["primary_soft"], foreground=c["accent"], bordercolor=c["border_strong"], relief="flat", font=(FONT_FAMILY, 9, "bold"))
        style.map("ResultAction.TButton", background=[("active", c["surface_tint"]), ("pressed", c["border"])], foreground=[("active", c["primary"])])
        style.configure("ResultHint.TLabel", background=c["surface"], foreground=c["accent"], font=(FONT_FAMILY, 9, "bold"))

        style.configure("TNotebook", background=c["surface"], borderwidth=0)
        style.configure("TNotebook.Tab", padding=(18, 9), background=c["surface"], foreground=c["muted"], borderwidth=0)
        style.map("TNotebook.Tab", background=[("selected", c["surface_alt"]), ("active", c["surface_alt"])], foreground=[("selected", c["primary"]), ("active", c["text"])])
        style.configure("Workspace.TNotebook", background=c["bg"], borderwidth=0, tabmargins=(0, 0, 0, 12))
        style.configure("Workspace.TNotebook.Tab", padding=(24, 11), background=c["bg"], foreground=c["muted"], borderwidth=0, font=(FONT_FAMILY, 11, "bold"))
        style.map("Workspace.TNotebook.Tab", background=[("selected", c["surface"]), ("active", c["surface_alt"])], foreground=[("selected", c["primary"]), ("active", c["accent"])])
        style.configure("TEntry", padding=(8, 6), fieldbackground=c["white"], foreground=c["text"], bordercolor=c["border"], lightcolor=c["border"], darkcolor=c["border"], relief="flat")
        style.map("TEntry", bordercolor=[("focus", c["primary"])], lightcolor=[("focus", c["primary"])], darkcolor=[("focus", c["primary"])])
        style.configure(
            "Scheme.TCombobox",
            padding=(12, 8),
            fieldbackground=c["surface_alt"],
            background=c["surface_alt"],
            foreground=c["accent"],
            bordercolor=c["border_strong"],
            lightcolor=c["border_strong"],
            darkcolor=c["border_strong"],
            arrowcolor=c["primary"],
            relief="flat",
            font=(FONT_FAMILY, 10, "bold"),
        )
        style.map(
            "Scheme.TCombobox",
            fieldbackground=[("readonly", c["surface_alt"]), ("focus", c["white"])],
            background=[("readonly", c["surface_alt"]), ("active", c["surface_tint"])],
            bordercolor=[("focus", c["primary"]), ("active", c["primary"])],
            foreground=[("readonly", c["accent"])],
        )
        style.configure("TCheckbutton", background=c["surface"], foreground=c["text"], padding=4, focuscolor=c["primary_soft"])
        style.map("TCheckbutton", background=[("active", c["surface"])], foreground=[("active", c["primary"])])
        style.configure("Tint.TCheckbutton", background=c["surface_alt"], foreground=c["text"], padding=4, focuscolor=c["primary_soft"])
        style.map("Tint.TCheckbutton", background=[("active", c["surface_alt"])], foreground=[("active", c["primary"])])
        style.configure("TSeparator", background=c["border"])
        style.configure("Treeview", background=c["white"], fieldbackground=c["white"], foreground=c["text"], borderwidth=0, rowheight=30)
        style.configure("Treeview.Heading", background=c["surface_alt"], foreground=c["text"], font=(FONT_FAMILY, 9, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", c["primary_soft"])], foreground=[("selected", c["text"])])
        style.configure("Settings.Treeview", background=c["white"], fieldbackground=c["white"], foreground=c["text"], borderwidth=0, rowheight=34)
        style.configure("Settings.Treeview.Heading", background=c["surface_tint"], foreground=c["accent"], font=(FONT_FAMILY, 9, "bold"), relief="flat")
        style.map("Settings.Treeview", background=[("selected", c["primary_soft"])], foreground=[("selected", c["primary"])])

        option = self.root.option_add
        for key, value in [
            ("*Listbox.font", (FONT_FAMILY, 10)),
            ("*Listbox.background", c["white"]),
            ("*Listbox.foreground", c["text"]),
            ("*Listbox.selectBackground", c["primary_soft"]),
            ("*Listbox.selectForeground", c["text"]),
            ("*Listbox.borderWidth", 0),
            ("*Listbox.highlightThickness", 1),
            ("*Listbox.highlightBackground", c["border"]),
            ("*Listbox.highlightColor", c["primary"]),
            ("*Listbox.relief", "flat"),
            ("*Text.font", (MONO_FAMILY, 10)),
            ("*Text.background", c["white"]),
            ("*Text.foreground", c["text"]),
            ("*Text.borderWidth", 0),
            ("*Text.highlightThickness", 1),
            ("*Text.highlightBackground", c["border"]),
            ("*Text.highlightColor", c["primary"]),
        ]:
            option(key, value)

    def _init_config_vars(self) -> None:
        page_fields = [
            "margin_top_cm",
            "margin_bottom_cm",
            "margin_left_cm",
            "margin_right_cm",
            "footer_distance_cm",
            "line_spacing_pt",
            "title_line_spacing_pt",
            "red_head_line_spacing_pt",
            "chars_per_line",
            "lines_per_page",
        ]
        self.page_vars = {name: tk.StringVar() for name in page_fields}
        self.format_vars = {
            name: tk.BooleanVar()
            for name in [
                "apply_page_setup",
                "apply_document_grid",
                "apply_styles",
                "add_page_number",
                "page_number_odd_even",
                "draw_red_separator",
                "draw_imprint_lines",
                "preserve_existing_bold_italic",
            ]
        }
        self.audit_vars = {
            name: tk.BooleanVar()
            for name in [
                "require_document_number_for_red_head",
                "require_signer_for_red_head",
                "require_main_send",
                "require_date",
                "check_page_layout",
                "check_document_grid",
                "check_unit_typography",
                "check_date_format",
                "check_attachment_format",
                "check_finalization_terms",
                "check_title_line_shape",
                "check_imprint_rules",
                "check_document_number_format",
                "check_attachment_layout",
                "check_front_matter_order",
            ]
        }
        self.style_vars = {
            name: (tk.StringVar(), tk.StringVar())
            for name in ["red_head", "title", "body", "h1", "h2", "h3", "copy_to", "page_number"]
        }
        self.generation_vars = {
            "add_red_head": tk.BooleanVar(value=False),
            "add_imprint": tk.BooleanVar(value=False),
            "red_head_title": tk.StringVar(),
            "document_number": tk.StringVar(),
            "copy_to": tk.StringVar(),
            "print_organization": tk.StringVar(),
            "print_date": tk.StringVar(),
            "meeting_number": tk.StringVar(),
            "meeting_organization": tk.StringVar(),
            "meeting_date": tk.StringVar(),
            "distribution": tk.StringVar(),
        }
        self._sync_form_from_config()

    def _build(self) -> None:
        outer = ttk.Frame(self.root, style="TFrame", padding=(24, 14, 24, 18))
        outer.pack(fill=tk.BOTH, expand=True)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)

        workspace = ttk.Notebook(outer, style="Workspace.TNotebook")
        workspace.grid(row=0, column=0, sticky="nsew")
        document_page = ttk.Frame(workspace, style="TFrame", padding=(0, 4, 0, 0))
        document_page.columnconfigure(0, weight=1)
        document_page.rowconfigure(0, weight=1)
        workspace.add(document_page, text="文档")
        workspace.add(self._table_page(workspace), text="表格")
        workspace.add(self._placeholder_tab(workspace, "其他功能正在开发中"), text="其他")

        main = ttk.Frame(document_page, style="TFrame")
        main.grid(row=0, column=0, sticky="nsew")
        main.columnconfigure(0, weight=2, uniform="main")
        main.columnconfigure(1, weight=3, uniform="main")
        main.rowconfigure(0, weight=1)

        left = ttk.Frame(main, style="Card.TFrame", padding=(20, 18, 20, 18))
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 16))
        left.columnconfigure(0, weight=1)
        left.rowconfigure(2, weight=1)
        left.rowconfigure(3, weight=1)

        queue_header = ttk.Frame(left, style="Card.TFrame")
        queue_header.grid(row=0, column=0, sticky="ew")
        queue_header.columnconfigure(0, weight=1)
        queue_tools = ttk.Frame(queue_header, style="Toolbar.TFrame")
        queue_tools.grid(row=0, column=0, sticky="ew")
        for column in range(2):
            queue_tools.columnconfigure(column, weight=1, uniform="queue_tools")
        for index, (text, command) in enumerate([
            ("添加文件", self.add_document_files),
            ("添加文件夹", self.add_document_folder),
            ("移除选中", self.remove_selected_documents),
            ("清空列表", self.clear_documents),
        ]):
            ttk.Button(queue_tools, text=text, style="Workbench.TButton", command=command).grid(
                row=index // 2,
                column=index % 2,
                sticky="ew",
                padx=(0, 6) if index % 2 == 0 else (6, 0),
                pady=(0, 6) if index < 2 else (0, 0),
            )

        list_wrap = ttk.Frame(left, style="Tint.TFrame", padding=1)
        list_wrap.grid(row=2, column=0, sticky="nsew", pady=(14, 0))
        list_wrap.columnconfigure(0, weight=1)
        list_wrap.rowconfigure(0, weight=1)
        self.doc_list = tk.Listbox(list_wrap, activestyle="none", exportselection=False)
        self.doc_list.grid(row=0, column=0, sticky="nsew")
        doc_scroll = ttk.Scrollbar(list_wrap, orient="vertical", command=self.doc_list.yview)
        doc_scroll.grid(row=0, column=1, sticky="ns")
        self.doc_list.configure(yscrollcommand=doc_scroll.set)
        self.doc_list_placeholder = tk.Label(
            self.doc_list,
            text="使用上方按钮导入文件或文件夹，可批量处理",
            bg=COLOR["white"],
            fg=COLOR["subtle"],
            font=(FONT_FAMILY, 9),
        )
        self.doc_list_placeholder.place(relx=0.5, rely=0.5, anchor="center")

        editor = ttk.Frame(left, style="Card.TFrame")
        editor.grid(row=3, column=0, sticky="nsew", pady=(14, 0))
        editor.columnconfigure(0, weight=1)
        editor.rowconfigure(0, weight=1)
        editor_wrap = ttk.Frame(editor, style="Tint.TFrame", padding=1)
        editor_wrap.grid(row=0, column=0, sticky="nsew")
        editor_wrap.columnconfigure(0, weight=1)
        editor_wrap.rowconfigure(0, weight=1)
        self.direct_text = tk.Text(editor_wrap, wrap="word", height=7, undo=True)
        self.direct_text.grid(row=0, column=0, sticky="nsew")
        editor_scroll = ttk.Scrollbar(editor_wrap, orient="vertical", command=self.direct_text.yview)
        editor_scroll.grid(row=0, column=1, sticky="ns")
        self.direct_text.configure(yscrollcommand=editor_scroll.set)
        _install_text_placeholder(self.direct_text, "点击这里输入文本内容，直接校对导出 Word 文档")

        actions = ttk.Frame(left, style="Card.TFrame")
        actions.grid(row=4, column=0, sticky="ew", pady=(16, 0))
        for column in range(2):
            actions.columnconfigure(column, weight=1, uniform="main_actions")
        for index, (text, command) in enumerate([
            ("开始校对", self.audit_documents),
            ("校对导出", self.format_documents),
            ("配置 AI", self.open_ai_settings),
            ("启用 AI", self.toggle_ai_review),
        ]):
            button_options = {"style": "Workbench.TButton", "command": command}
            if text == "启用 AI":
                button_options["textvariable"] = self.ai_toggle_text
            else:
                button_options["text"] = text
            ttk.Button(actions, **button_options).grid(
                row=index // 2,
                column=index % 2,
                sticky="ew",
                padx=(0, 6) if index % 2 == 0 else (6, 0),
                pady=(0, 8) if index < 2 else (0, 0),
            )

        right = ttk.Frame(main, style="Card.TFrame", padding=(1, 1, 1, 1))
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)
        notebook = ttk.Notebook(right)
        notebook.grid(row=0, column=0, sticky="nsew")
        notebook.add(self._results_tab(notebook), text="校对结果")
        notebook.add(self._audit_tab(notebook), text="规则格式")

        status = ttk.Frame(document_page, style="Tint.TFrame", padding=(12, 8))
        status.grid(row=1, column=0, sticky="ew", pady=(14, 0))
        ttk.Label(status, textvariable=self.status_text, style="Status.TLabel").pack(anchor="w")

    @staticmethod
    def _placeholder_tab(parent: ttk.Notebook, message: str) -> ttk.Frame:
        frame = ttk.Frame(parent, style="Card.TFrame", padding=32)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        panel = ttk.Frame(frame, style="Tint.TFrame", padding=(28, 24))
        panel.grid(row=0, column=0)
        ttk.Label(panel, text=message, style="Tint.TLabel", font=(FONT_FAMILY, 12, "bold")).pack()
        return frame

    def _table_page(self, parent: ttk.Notebook) -> ttk.Frame:
        page = ttk.Frame(parent, style="TFrame", padding=(0, 4, 0, 0))
        page.columnconfigure(0, weight=1)
        page.rowconfigure(0, weight=1)

        modes = ttk.Notebook(page)
        modes.grid(row=0, column=0, sticky="nsew")
        merge_page = ttk.Frame(modes, style="TFrame", padding=(0, 0, 0, 0))
        split_page = self._placeholder_tab(modes, "拆分表格功能正在开发中")
        inspect_page = self._placeholder_tab(modes, "表格检查功能将独立整理")
        modes.add(merge_page, text="汇总填报")
        modes.add(split_page, text="拆分表格")
        modes.add(inspect_page, text="表格检查")

        merge_page.columnconfigure(0, weight=1)
        merge_page.rowconfigure(0, weight=1)
        main = ttk.Frame(merge_page, style="TFrame")
        main.grid(row=0, column=0, sticky="nsew", pady=(12, 0))
        main.columnconfigure(0, weight=4, uniform="table")
        main.columnconfigure(1, weight=5, uniform="table")
        main.columnconfigure(2, weight=4, uniform="table")
        main.rowconfigure(0, weight=1)

        source_panel = ttk.Frame(main, style="Card.TFrame", padding=(18, 16, 18, 16))
        source_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        source_panel.columnconfigure(0, weight=1)
        source_panel.rowconfigure(1, weight=1)

        list_wrap = ttk.Frame(source_panel, style="Tint.TFrame", padding=1)
        list_wrap.grid(row=1, column=0, sticky="nsew", pady=(0, 0))
        list_wrap.columnconfigure(0, weight=1)
        list_wrap.rowconfigure(0, weight=1)
        self.table_list = tk.Listbox(list_wrap, activestyle="none", exportselection=False)
        self.table_list.grid(row=0, column=0, sticky="nsew")
        table_scroll = ttk.Scrollbar(list_wrap, orient="vertical", command=self.table_list.yview)
        table_scroll.grid(row=0, column=1, sticky="ns")
        self.table_list.configure(yscrollcommand=table_scroll.set)
        self.table_list_placeholder = tk.Label(
            self.table_list,
            text="导入主表和副表文件",
            bg=COLOR["white"],
            fg=COLOR["subtle"],
            font=(FONT_FAMILY, 9),
        )
        self.table_list_placeholder.place(relx=0.5, rely=0.5, anchor="center")
        self.table_list.bind("<<ListboxSelect>>", self._on_table_source_selected)

        source_tools = ttk.Frame(source_panel, style="Toolbar.TFrame")
        source_tools.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        for column in range(2):
            source_tools.columnconfigure(column, weight=1, uniform="source_tools")
        for index, (text, command) in enumerate([
            ("添加文件", self.add_table_files),
            ("添加文件夹", self.add_table_folder),
            ("移除选中", self.remove_selected_tables),
            ("清空列表", self.clear_tables),
        ]):
            ttk.Button(source_tools, text=text, style="Workbench.TButton", command=command).grid(
                row=index // 2,
                column=index % 2,
                sticky="ew",
                padx=(0, 6) if index % 2 == 0 else (6, 0),
                pady=(0, 6) if index < 2 else (0, 0),
            )

        fuzzy = ttk.Frame(source_panel, style="Toolbar.TFrame")
        fuzzy.grid(row=2, column=0, sticky="ew", pady=(12, 0))
        fuzzy.columnconfigure(1, weight=1)
        ttk.Checkbutton(fuzzy, text="模糊匹配", variable=self.table_fuzzy_enabled).grid(row=0, column=0, sticky="w", padx=(0, 10))
        ttk.Scale(
            fuzzy,
            from_=60,
            to=100,
            variable=self.table_fuzzy_threshold,
            orient="horizontal",
            command=lambda value: self.table_fuzzy_threshold.set(int(float(value))),
        ).grid(row=0, column=1, sticky="ew", padx=(0, 10))
        ttk.Label(fuzzy, textvariable=self.table_fuzzy_threshold, style="Card.TLabel", width=4).grid(row=0, column=2, sticky="e")

        rules_panel = ttk.Frame(main, style="Card.TFrame", padding=(18, 16, 18, 16))
        rules_panel.grid(row=0, column=1, sticky="nsew", padx=(0, 12))
        rules_panel.columnconfigure(0, weight=1)
        rules_panel.rowconfigure(2, weight=1)

        role = ttk.Frame(rules_panel, style="Toolbar.TFrame")
        role.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        for column in range(2):
            role.columnconfigure(column, weight=1, uniform="table_top_actions")
        for index, (text, command) in enumerate([
            ("设为主表", self.use_selected_table_as_master),
            ("一键汇总", self.merge_same_layout_tables),
            ("添加副表", self.add_table_files),
            ("删除副表", self.remove_selected_table_setting),
        ]):
            ttk.Button(role, text=text, style="Workbench.TButton", command=command).grid(
                row=index // 2,
                column=index % 2,
                sticky="ew",
                padx=(0, 6) if index % 2 == 0 else (6, 0),
                pady=(0, 6) if index < 2 else (0, 0),
            )

        settings = ttk.Frame(rules_panel, style="Tint.TFrame", padding=(1, 1))
        settings.grid(row=2, column=0, sticky="nsew")
        settings.columnconfigure(0, weight=1)
        settings.rowconfigure(1, weight=1)

        editor = ttk.Frame(settings, style="Card.TFrame", padding=(10, 8))
        editor.grid(row=0, column=0, sticky="ew")
        for column in range(3):
            editor.columnconfigure(column, weight=1, uniform="table_editor")
        self.table_selected_source_sheet_box = ttk.Combobox(editor, textvariable=self.table_selected_source_sheet)
        self.table_selected_source_sheet_box.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self.table_selected_source_key_box = ttk.Combobox(editor, textvariable=self.table_selected_source_key_column)
        self.table_selected_source_key_box.grid(row=0, column=1, sticky="ew", padx=(0, 8))
        self.table_selected_source_value_box = ttk.Combobox(editor, textvariable=self.table_selected_source_value_column)
        self.table_selected_source_value_box.grid(row=0, column=2, sticky="ew")
        for box in [self.table_selected_source_sheet_box, self.table_selected_source_key_box, self.table_selected_source_value_box]:
            box.bind("<<ComboboxSelected>>", lambda _e: self._on_selected_table_setting_changed())
            box.bind("<FocusOut>", lambda _e: self._save_selected_table_settings(silent=True))

        self.table_rules_tree = ttk.Treeview(
            settings,
            columns=("role", "file", "sheet", "key", "value"),
            show="headings",
            height=9,
            style="Settings.Treeview",
        )
        for col, label, width in [
            ("role", "类型", 54),
            ("file", "文件", 130),
            ("sheet", "工作表", 90),
            ("key", "匹配列", 90),
            ("value", "取值列", 90),
        ]:
            self.table_rules_tree.heading(col, text=label, anchor="center")
            self.table_rules_tree.column(col, width=width, minwidth=50, stretch=True, anchor="center")
        self.table_rules_tree.grid(row=1, column=0, sticky="nsew")
        self.table_rules_tree.bind("<<TreeviewSelect>>", self._on_table_settings_selected)
        rules_scroll = ttk.Scrollbar(settings, orient="vertical", command=self.table_rules_tree.yview)
        rules_scroll.grid(row=1, column=1, sticky="ns")
        self.table_rules_tree.configure(yscrollcommand=rules_scroll.set)

        actions = ttk.Frame(settings, style="Toolbar.TFrame")
        actions.grid(row=2, column=0, sticky="ew")
        for col in range(2):
            actions.columnconfigure(col, weight=1, uniform="table_actions")
        for index, (text, command) in enumerate([
            ("预览汇总", self.preview_table_merge),
            ("导出汇总", self.export_table_merge),
        ]):
            ttk.Button(actions, text=text, style="Workbench.TButton", command=command).grid(
                row=0,
                column=index,
                sticky="ew",
                padx=(0, 6) if index % 2 == 0 else (6, 0),
            )

        right = ttk.Frame(main, style="Card.TFrame", padding=(18, 16, 18, 16))
        right.grid(row=0, column=2, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)

        preview_wrap = ttk.Frame(right, style="Tint.TFrame", padding=1)
        preview_wrap.grid(row=0, column=0, sticky="nsew")
        preview_wrap.columnconfigure(0, weight=1)
        preview_wrap.rowconfigure(0, weight=1)
        self.table_preview_canvas = tk.Canvas(preview_wrap, bg=COLOR["white"], highlightthickness=0)
        self.table_preview_canvas.grid(row=0, column=0, sticky="nsew")
        self.table_preview_frame = ttk.Frame(self.table_preview_canvas, style="Card.TFrame")
        preview_window = self.table_preview_canvas.create_window((0, 0), window=self.table_preview_frame, anchor="nw")
        tree_scroll = ttk.Scrollbar(preview_wrap, orient="vertical", command=self.table_preview_canvas.yview)
        tree_scroll.grid(row=0, column=1, sticky="ns")
        tree_xscroll = ttk.Scrollbar(preview_wrap, orient="horizontal", command=self.table_preview_canvas.xview)
        tree_xscroll.grid(row=1, column=0, sticky="ew")
        self.table_preview_canvas.configure(yscrollcommand=tree_scroll.set, xscrollcommand=tree_xscroll.set)
        self.table_preview_frame.bind("<Configure>", lambda _e: self.table_preview_canvas.configure(scrollregion=self.table_preview_canvas.bbox("all")))
        self.table_preview_canvas.bind("<Configure>", lambda e: self.table_preview_canvas.itemconfigure(preview_window, height=max(1, e.height)))
        self.table_tree_placeholder = tk.Label(
            self.table_preview_canvas,
            text="选择主表或副表后，完整表格内容将在这里预览",
            bg=COLOR["white"],
            fg=COLOR["subtle"],
            font=(FONT_FAMILY, 9),
        )
        self.table_tree_placeholder.place(relx=0.5, rely=0.5, anchor="center")

        status = ttk.Frame(page, style="Tint.TFrame", padding=(12, 8))
        status.grid(row=1, column=0, sticky="ew", pady=(14, 0))
        ttk.Label(status, textvariable=self.status_text, style="Status.TLabel").pack(anchor="w")

        return page

    def _results_tab(self, parent: ttk.Notebook) -> ttk.Frame:
        frame = ttk.Frame(parent, style="Card.TFrame", padding=(18, 16, 18, 16))
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        frame.rowconfigure(2, weight=1)

        columns = ("file", "type", "source", "summary")
        self.result_tree = ttk.Treeview(frame, columns=columns, show="headings", height=10)
        headings = {"file": "文件", "type": "错误类型", "source": "来源", "summary": "摘要"}
        widths = {"file": 150, "type": 110, "source": 90, "summary": 460}
        for col in columns:
            self.result_tree.heading(col, text=headings[col])
            self.result_tree.column(col, width=widths[col], minwidth=70, stretch=col == "summary")
        self.result_tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll = ttk.Scrollbar(frame, orient="vertical", command=self.result_tree.yview)
        tree_scroll.grid(row=0, column=1, sticky="ns")
        self.result_tree.configure(yscrollcommand=tree_scroll.set)
        self.result_tree.bind("<<TreeviewSelect>>", self._show_selected_result)
        self.result_tree_placeholder = tk.Label(
            self.result_tree,
            text="完成校对后，结果将在这里显示",
            bg=COLOR["white"],
            fg=COLOR["subtle"],
            font=(FONT_FAMILY, 9),
        )
        self.result_tree_placeholder.place(relx=0.5, rely=0.5, anchor="center")

        footer = ttk.Frame(frame, style="Card.TFrame")
        footer.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        ttk.Button(footer, text="清空结果", style="ResultAction.TButton", command=self._clear_results).pack(side=tk.LEFT)
        ttk.Label(footer, text="选中结果后，下方显示完整内容与处理建议。", style="ResultHint.TLabel").pack(side=tk.LEFT, padx=(12, 0))

        detail = ttk.Frame(frame, style="Tint.TFrame", padding=(12, 10, 12, 10))
        detail.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=(10, 0))
        detail.columnconfigure(0, weight=1)
        detail.rowconfigure(1, weight=1)
        ttk.Label(detail, text="详情", style="Status.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.result_detail_text = tk.Text(detail, wrap="word", height=8)
        self.result_detail_text.grid(row=1, column=0, sticky="nsew")
        detail_scroll = ttk.Scrollbar(detail, orient="vertical", command=self.result_detail_text.yview)
        detail_scroll.grid(row=1, column=1, sticky="ns")
        self.result_detail_text.configure(yscrollcommand=detail_scroll.set, state="disabled")
        self._set_result_detail("选中上方任意一条校对结果，可查看完整内容。")
        return frame

    def _audit_tab(self, parent: ttk.Notebook) -> ttk.Frame:
        wrap, body = _make_scrollable(parent)
        for column in range(5):
            body.columnconfigure(column, weight=1, minsize=150)

        def section(row: int, title: str, subtitle: str = "") -> int:
            if row:
                ttk.Separator(body, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=5, sticky="ew", pady=18)
                row += 1
            ttk.Label(body, text=title, style="Section.TLabel").grid(row=row, column=0, columnspan=5, sticky="w")
            if subtitle:
                ttk.Label(body, text=subtitle, style="Muted.TLabel").grid(row=row + 1, column=0, columnspan=5, sticky="w", pady=(4, 12))
                return row + 2
            return row + 1

        def feature_button(parent, text: str, variables: list[tk.BooleanVar], row: int, column: int, padx=(0, 12), pady=4) -> tk.Button:
            def is_enabled() -> bool:
                return all(variable.get() for variable in variables)

            def refresh() -> None:
                if is_enabled():
                    button.configure(
                        bg=COLOR["primary"],
                        fg=COLOR["white"],
                        activebackground=COLOR["primary_hover"],
                        activeforeground=COLOR["white"],
                    )
                else:
                    button.configure(
                        bg=COLOR["surface"],
                        fg=COLOR["accent"],
                        activebackground=COLOR["surface_tint"],
                        activeforeground=COLOR["accent"],
                    )

            def toggle() -> None:
                enabled = not is_enabled()
                for variable in variables:
                    variable.set(enabled)
                refresh()

            button = tk.Button(
                parent,
                text=text,
                command=toggle,
                relief="flat",
                bd=0,
                padx=14,
                pady=8,
                cursor="hand2",
                font=(FONT_FAMILY, 10),
                highlightthickness=1,
                highlightbackground=COLOR["border"],
                highlightcolor=COLOR["primary"],
            )
            button.grid(row=row, column=column, sticky="ew", padx=padx, pady=pady)
            refresh()
            for variable in variables:
                variable.trace_add("write", lambda *_args: refresh())
            return button

        deferred_audit_options = (
            "require_signer_for_red_head",
            "require_main_send",
            "require_date",
            "check_finalization_terms",
            "check_title_line_shape",
        )
        for key in deferred_audit_options:
            self.audit_vars[key].set(False)

        scheme_notes = {
            "普通公文": "适合常规通知、报告等文本，使用通用 A4 页边距和正文格式。",
            "红头文件": "适合正式红头文件，重点处理版头、发文字号、红线和页码。",
            "红头文件（函）": "适合函格式红头文件：正文锚定文本框版头、双红线、函号右对齐、首页无页码。",
            "红头文件（会议纪要）": "适合会议纪要：专用版头、期号、编发信息、出席人员和分送版记。",
        }

        def apply_scheme(name: str) -> None:
            if name in self.config_profiles:
                ai_options = deepcopy(self.config.ai_review)
                self.config = deepcopy(self.config_profiles[name])
                self.config.ai_review = ai_options
                self._sync_form_from_config()
                self.scheme_var.set(_profile_label(self.config.audit.profile))
                scheme_note.set(f"已载入自定义配置“{name}”。")
                self._refresh_generation_panel()
                return
            presets = {
                "普通公文": {
                    "profile": "standard",
                    "page": {
                        "margin_top_cm": 3.7,
                        "margin_bottom_cm": 3.5,
                        "margin_left_cm": 2.8,
                        "margin_right_cm": 2.6,
                        "footer_distance_cm": 2.5,
                        "chars_per_line": 28,
                        "lines_per_page": 22,
                    },
                    "format": {
                        "apply_page_setup": True,
                        "apply_document_grid": True,
                        "apply_styles": True,
                        "add_page_number": True,
                        "page_number_odd_even": True,
                        "draw_red_separator": False,
                        "preserve_existing_bold_italic": True,
                    },
                    "audit": {
                        "require_document_number_for_red_head": True,
                        "require_signer_for_red_head": False,
                        "require_main_send": False,
                        "require_date": True,
                        "check_page_layout": True,
                        "check_document_grid": True,
                        "check_unit_typography": True,
                        "check_date_format": True,
                        "check_attachment_format": True,
                        "check_finalization_terms": True,
                    },
                },
                "红头文件": {
                    "profile": "red_head",
                    "page": {
                        "margin_top_cm": 3.7,
                        "margin_bottom_cm": 3.5,
                        "margin_left_cm": 2.8,
                        "margin_right_cm": 2.6,
                        "footer_distance_cm": 2.5,
                        "chars_per_line": 28,
                        "lines_per_page": 22,
                    },
                    "format": {
                        "apply_page_setup": True,
                        "apply_document_grid": True,
                        "apply_styles": True,
                        "add_page_number": True,
                        "page_number_odd_even": True,
                        "draw_red_separator": True,
                        "preserve_existing_bold_italic": True,
                    },
                    "audit": {
                        "require_document_number_for_red_head": True,
                        "require_signer_for_red_head": False,
                        "require_main_send": False,
                        "require_date": True,
                        "check_page_layout": True,
                        "check_document_grid": True,
                        "check_unit_typography": True,
                        "check_date_format": True,
                        "check_attachment_format": True,
                        "check_finalization_terms": True,
                    },
                },
                "红头文件（函）": {
                    "profile": "letter_head",
                    "page": {
                        "margin_top_cm": 3.7,
                        "margin_bottom_cm": 2.5,
                        "margin_left_cm": 2.8,
                        "margin_right_cm": 2.6,
                        "footer_distance_cm": 2.5,
                        "chars_per_line": 28,
                        "lines_per_page": 22,
                    },
                    "format": {
                        "apply_page_setup": True,
                        "apply_document_grid": True,
                        "apply_styles": True,
                        "add_page_number": True,
                        "page_number_odd_even": True,
                        "draw_red_separator": True,
                        "preserve_existing_bold_italic": True,
                    },
                    "audit": {
                        "require_document_number_for_red_head": True,
                        "require_signer_for_red_head": False,
                        "require_main_send": False,
                        "require_date": True,
                        "check_page_layout": True,
                        "check_document_grid": True,
                        "check_unit_typography": True,
                        "check_date_format": True,
                        "check_attachment_format": True,
                        "check_finalization_terms": True,
                    },
                },
                "红头文件（会议纪要）": {
                    "profile": "meeting_minutes",
                    "page": {
                        "margin_top_cm": 3.7,
                        "margin_bottom_cm": 3.5,
                        "margin_left_cm": 2.8,
                        "margin_right_cm": 2.6,
                        "footer_distance_cm": 2.5,
                        "chars_per_line": 28,
                        "lines_per_page": 22,
                    },
                    "format": {
                        "apply_page_setup": True,
                        "apply_document_grid": True,
                        "apply_styles": True,
                        "add_page_number": True,
                        "page_number_odd_even": True,
                        "draw_red_separator": True,
                        "preserve_existing_bold_italic": True,
                    },
                    "audit": {
                        "require_document_number_for_red_head": False,
                        "require_signer_for_red_head": False,
                        "require_main_send": False,
                        "require_date": False,
                        "check_page_layout": True,
                        "check_document_grid": True,
                        "check_unit_typography": True,
                        "check_date_format": True,
                        "check_attachment_format": True,
                        "check_finalization_terms": True,
                    },
                },
            }
            preset = presets[name]
            self.config.audit.profile = preset["profile"]
            default_config = OfficeToolConfig()
            for key, value in preset.get("page", {}).items():
                self.page_vars[key].set(str(value))
            for key, value in preset["format"].items():
                self.format_vars[key].set(value)
            for key, variable in self.format_vars.items():
                if key not in preset["format"]:
                    variable.set(bool(getattr(default_config.format, key)))
            for key, value in preset["audit"].items():
                self.audit_vars[key].set(value)
            for key, variable in self.audit_vars.items():
                if key not in preset["audit"]:
                    variable.set(bool(getattr(default_config.audit, key)))
            for key in deferred_audit_options:
                self.audit_vars[key].set(False)
            scheme_note.set(scheme_notes[name])
            self._refresh_generation_panel()

        self._apply_scheme_callback = apply_scheme

        def restore_defaults() -> None:
            self.config = _default_gui_config(self.config.ai_review)
            self._sync_form_from_config()
            for key in deferred_audit_options:
                self.audit_vars[key].set(False)
            self.scheme_var.set("普通公文")
            scheme_note.set("已恢复默认公文规则参数。")
            self._refresh_generation_panel()

        row = section(0, "规则方案", "先选择常用场景，再按需要微调检查和修复项目。")
        scheme_note = tk.StringVar(value=scheme_notes["普通公文"])
        scheme_panel = ttk.Frame(body, style="Tint.TFrame", padding=(12, 9, 12, 9))
        scheme_panel.grid(row=row, column=0, columnspan=2, sticky="ew", padx=(0, 12), pady=(0, 8))
        scheme_panel.columnconfigure(1, weight=1)
        ttk.Label(scheme_panel, text="当前方案", style="Tint.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.scheme_box = ttk.Combobox(
            scheme_panel,
            textvariable=self.scheme_var,
            values=BUILTIN_SCHEMES,
            state="readonly",
            width=18,
            style="Scheme.TCombobox",
        )
        self.scheme_box.grid(row=0, column=1, sticky="ew")
        ttk.Button(body, text="应用方案", style="Primary.TButton", command=lambda: apply_scheme(self.scheme_var.get())).grid(row=row, column=2, sticky="ew", padx=(0, 12), pady=(0, 8))
        ttk.Button(body, text="恢复默认", style="Config.TButton", command=restore_defaults).grid(row=row, column=3, sticky="ew", pady=(0, 8))
        config_button = ttk.Button(body, text="自定义配置 ▾", style="Config.TButton")
        self.config_profile_button = config_button
        config_button.grid(row=row, column=4, sticky="ew", padx=(12, 0), pady=(0, 8))
        config_button.configure(command=lambda: self.open_config_profile_popup(config_button))
        ttk.Label(body, textvariable=scheme_note, style="Muted.TLabel", wraplength=900).grid(row=row + 1, column=0, columnspan=5, sticky="w")
        def scheme_selected(_event=None) -> None:
            name = self.scheme_var.get()
            scheme_note.set(scheme_notes.get(name, f"自定义配置“{name}”，点击应用方案后生效。"))
            self._refresh_generation_panel()

        self.scheme_box.bind("<<ComboboxSelected>>", scheme_selected)

        generation_frame = ttk.Frame(body, style="Tint.TFrame", padding=(12, 10, 12, 10))
        generation_frame.grid(row=row + 2, column=0, columnspan=5, sticky="ew", pady=(10, 0))
        generation_frame.columnconfigure(0, weight=1)
        generation_toggle_frame = ttk.Frame(generation_frame, style="Tint.TFrame")
        generation_toggle_frame.grid(row=0, column=0, sticky="ew")
        generation_toggle_frame.columnconfigure(0, weight=1)
        generation_toggle_frame.columnconfigure(1, weight=1)
        add_red_head_button = feature_button(
            generation_toggle_frame,
            "添加红头",
            [self.generation_vars["add_red_head"]],
            0,
            0,
            padx=(0, 8),
            pady=0,
        )
        add_imprint_button = feature_button(
            generation_toggle_frame,
            "添加版记",
            [self.generation_vars["add_imprint"]],
            0,
            1,
            padx=(8, 0),
            pady=0,
        )
        generation_hint = ttk.Label(
            generation_frame,
            text="发文字号提供可编辑示例；必填项为空时不导出。函号可删除留空，普通红头的抄送可留空并生成简版版记。",
            style="TintMuted.TLabel",
            wraplength=900,
        )
        generation_hint.grid(row=1, column=0, sticky="w", pady=(8, 0))
        generation_fields = ttk.Frame(generation_frame, style="Tint.TFrame")
        generation_fields.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        generation_fields.columnconfigure(1, weight=1)
        generation_fields.columnconfigure(3, weight=1)

        def selected_profile() -> str:
            name = self.scheme_var.get()
            return {
                "普通公文": "standard",
                "红头文件": "red_head",
                "红头文件（函）": "letter_head",
                "红头文件（会议纪要）": "meeting_minutes",
            }.get(name, self.config.audit.profile)

        def refresh_generation_panel() -> None:
            profile = selected_profile()
            if profile not in {"red_head", "letter_head", "meeting_minutes"}:
                generation_frame.grid_remove()
                return
            generation_frame.grid()
            if profile == "letter_head":
                add_imprint_button.grid_remove()
                if self.generation_vars["add_imprint"].get():
                    self.generation_vars["add_imprint"].set(False)
                    return
            else:
                add_imprint_button.grid()
            for child in generation_fields.winfo_children():
                child.destroy()

            fields: list[tuple[str, str]] = []
            if self.generation_vars["add_red_head"].get():
                if profile == "meeting_minutes":
                    fields.extend([
                        ("会议期号", "meeting_number"),
                        ("编发单位", "meeting_organization"),
                        ("编发日期", "meeting_date"),
                    ])
                else:
                    fields.extend([("红头名称", "red_head_title"), ("发文字号", "document_number")])
            if self.generation_vars["add_imprint"].get():
                if profile == "meeting_minutes":
                    fields.append(("分送内容", "distribution"))
                else:
                    fields.extend([
                        ("抄送内容", "copy_to"),
                        ("印发单位", "print_organization"),
                        ("印发日期", "print_date"),
                    ])
            for index, (label, key) in enumerate(fields):
                row_index, pair = divmod(index, 2)
                column = pair * 2
                ttk.Label(generation_fields, text=label, style="Tint.TLabel").grid(row=row_index, column=column, sticky="w", padx=(0, 8), pady=4)
                ttk.Entry(generation_fields, textvariable=self.generation_vars[key]).grid(row=row_index, column=column + 1, sticky="ew", padx=(0, 16) if pair == 0 else (0, 0), pady=4)
            if fields:
                generation_fields.grid()
            else:
                generation_fields.grid_remove()

        self._refresh_generation_panel = refresh_generation_panel
        self.generation_vars["add_red_head"].trace_add("write", lambda *_args: refresh_generation_panel())
        self.generation_vars["add_imprint"].trace_add("write", lambda *_args: refresh_generation_panel())
        refresh_generation_panel()

        row = section(row + 3, "处理项目", "启用某项后，“开始校对”会检查，“校对导出”会按同一项修复。")
        process_frame = ttk.Frame(body, style="Card.TFrame")
        process_frame.grid(row=row, column=0, columnspan=5, sticky="ew")
        for column in range(2):
            process_frame.columnconfigure(column, weight=1)
        process_items = [
            ("页面版式", [self.audit_vars["check_page_layout"], self.format_vars["apply_page_setup"]]),
            ("文档网格", [self.audit_vars["check_document_grid"], self.format_vars["apply_document_grid"]]),
            ("字体字号", [self.audit_vars["check_unit_typography"], self.format_vars["apply_styles"]]),
            ("页码格式", [self.format_vars["add_page_number"]]),
        ]
        for index, (label, variables) in enumerate(process_items):
            feature_button(
                process_frame,
                label,
                variables,
                index // 2,
                index % 2,
                padx=(0, 8) if index % 2 == 0 else (8, 0),
            )

        row = row + 2 + (len(process_items) + 1) // 2
        custom_visible = tk.BooleanVar(value=False)
        custom_button_text = tk.StringVar(value="展开自定义页面和字体")
        custom_frame = ttk.Frame(body, style="Card.TFrame")

        def toggle_custom_settings() -> None:
            custom_visible.set(not custom_visible.get())
            if custom_visible.get():
                custom_frame.grid()
                custom_button_text.set("收起自定义页面和字体")
            else:
                custom_frame.grid_remove()
                custom_button_text.set("展开自定义页面和字体")

        ttk.Button(body, textvariable=custom_button_text, style="Secondary.TButton", command=toggle_custom_settings).grid(
            row=row, column=0, columnspan=5, sticky="ew", pady=(18, 8)
        )
        custom_frame.grid(row=row + 1, column=0, columnspan=5, sticky="ew")
        custom_frame.grid_remove()
        for column in range(4):
            custom_frame.columnconfigure(column, weight=1, minsize=150)

        ttk.Label(custom_frame, text="页面版式", style="Section.TLabel").grid(row=0, column=0, columnspan=4, sticky="w")
        ttk.Label(custom_frame, text="一般按规则方案自动设置，确需微调时再展开修改。", style="Muted.TLabel").grid(
            row=1, column=0, columnspan=4, sticky="w", pady=(4, 12)
        )
        page_fields = [
            ("上边距 (cm)", "margin_top_cm"),
            ("下边距 (cm)", "margin_bottom_cm"),
            ("左边距 (cm)", "margin_left_cm"),
            ("右边距 (cm)", "margin_right_cm"),
            ("页脚距 (cm)", "footer_distance_cm"),
            ("正文行距 (pt)", "line_spacing_pt"),
            ("标题行距 (pt)", "title_line_spacing_pt"),
            ("红头行距 (pt)", "red_head_line_spacing_pt"),
            ("每行字数", "chars_per_line"),
            ("每页行数", "lines_per_page"),
        ]
        for index, (label, key) in enumerate(page_fields):
            current = 2 + index // 2
            col = (index % 2) * 2
            ttk.Label(custom_frame, text=label, style="Card.TLabel").grid(row=current, column=col, sticky="w", padx=(0, 8), pady=5)
            ttk.Entry(custom_frame, width=12, textvariable=self.page_vars[key]).grid(row=current, column=col + 1, sticky="ew", padx=(0, 20), pady=5)

        style_row = 2 + (len(page_fields) + 1) // 2
        ttk.Separator(custom_frame, orient=tk.HORIZONTAL).grid(row=style_row, column=0, columnspan=4, sticky="ew", pady=18)
        style_row += 1
        ttk.Label(custom_frame, text="字体样式", style="Section.TLabel").grid(row=style_row, column=0, columnspan=4, sticky="w")
        ttk.Label(custom_frame, text="按单位常用公文格式设置版头、标题、正文、版记和页码字体。", style="Muted.TLabel").grid(
            row=style_row + 1, column=0, columnspan=4, sticky="w", pady=(4, 12)
        )
        font_choices = [
            "方正小标宋简体",
            "华文中宋",
            "仿宋_GB2312",
            "仿宋",
            "楷体_GB2312",
            "楷体",
            "黑体",
            "宋体",
            "微软雅黑",
            "Microsoft YaHei UI",
        ]
        size_choices = list(FONT_SIZE_LABELS) + ["42", "36", "26", "24", "22", "18", "16", "15", "14", "12", "10.5", "9"]
        style_items = [
            ("red_head", "红头版头"),
            ("title", "主标题"),
            ("body", "正文"),
            ("h1", "一级标题"),
            ("h2", "二级标题"),
            ("h3", "三级标题"),
            ("copy_to", "抄送/印发"),
            ("page_number", "页码"),
        ]

        def add_style_row(container, index: int, key: str, label: str) -> None:
            font_var, size_var = self.style_vars[key]
            container.columnconfigure(1, weight=1)
            ttk.Label(container, text=label, style="Card.TLabel").grid(row=index, column=0, sticky="w", pady=5, padx=(0, 10))
            ttk.Combobox(container, textvariable=font_var, values=font_choices).grid(row=index, column=1, sticky="ew", padx=(0, 12), pady=5)
            ttk.Combobox(container, width=8, textvariable=size_var, values=size_choices).grid(row=index, column=2, sticky="ew", pady=5)

        style_frame = ttk.Frame(custom_frame, style="Card.TFrame")
        style_frame.grid(row=style_row + 2, column=0, columnspan=4, sticky="ew")
        for index, (key, label) in enumerate(style_items):
            add_style_row(style_frame, index, key, label)
        return wrap

    def add_document_files(self) -> None:
        paths = filedialog.askopenfilenames(title="选择公文文件", filetypes=[("支持的文档", "*.doc *.docx *.txt *.md"), ("所有文件", "*.*")])
        self._add_document_paths(paths)

    def choose_table_master(self) -> None:
        path = filedialog.askopenfilename(title="选择主表", filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")])
        if not path:
            return
        self.table_master_path = Path(path).resolve()
        self.table_master_file.set(self.table_master_path.name)
        self.table_output_file.set(str(self.table_master_path.with_name(f"{self.table_master_path.stem}_汇总结果.xlsx")))
        self._load_table_workbook_info(self.table_master_path)
        self._refresh_master_table_choices()
        self._refresh_table_source_list_labels()
        self._preview_table_path(self.table_master_path, self.table_master_sheet.get().strip() or None)
        self.status_text.set("已选择主表")

    def add_table_files(self) -> None:
        paths = filedialog.askopenfilenames(title="选择表格文件", filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")])
        self._add_table_paths(paths)

    def add_table_folder(self) -> None:
        folder = filedialog.askdirectory(title="选择副表文件夹")
        if folder:
            self._add_table_paths([folder])

    def _add_table_paths(self, paths) -> None:
        for raw in paths:
            path = Path(raw).resolve()
            if path not in self.table_paths:
                self.table_paths.append(path)
                self._load_table_workbook_info(path)
        self._refresh_table_placeholder()
        self._refresh_default_source_choices()
        self._refresh_table_source_list_labels()
        if paths:
            self.status_text.set(f"表格队列中 {len(self.table_paths)} 个路径")

    def use_selected_table_as_master(self) -> None:
        selection = self.table_list.curselection()
        if selection:
            path = self.table_paths[selection[0]].resolve()
        else:
            selected_setting = self._selected_table_setting_path()
            path = selected_setting.resolve() if selected_setting is not None else None
        if path is None:
            messagebox.showwarning("选择主表", "请先选中一张表。")
            return
        self.table_master_path = path
        self.table_master_file.set(path.name)
        self.table_output_file.set(str(path.with_name(f"{path.stem}_汇总结果.xlsx")))
        self._load_table_workbook_info(path)
        self._refresh_master_table_choices()
        self._refresh_default_source_choices()
        self._refresh_table_source_list_labels()
        self._select_table_setting(path)
        self._load_selected_table_settings(path)
        self._preview_table_path(path, self.table_master_sheet.get().strip() or None)
        self.status_text.set("已将选中表格设为主表")

    def remove_selected_tables(self) -> None:
        for index in reversed(list(self.table_list.curselection())):
            path = self.table_paths[index]
            self.table_list.delete(index)
            self.table_source_overrides.pop(path, None)
            if self.table_master_path is not None and path.resolve() == self.table_master_path.resolve():
                self.table_master_path = None
                self.table_master_file.set("")
                self.table_output_file.set("")
            del self.table_paths[index]
        self.table_selected_source_path = None
        self._clear_selected_source_form()
        self._refresh_table_placeholder()
        self._refresh_default_source_choices()
        self.status_text.set(f"表格队列中 {len(self.table_paths)} 个路径")

    def clear_tables(self) -> None:
        self.table_list.delete(0, tk.END)
        self.table_paths.clear()
        self.table_source_overrides.clear()
        self.table_selected_source_path = None
        self._clear_selected_source_form()
        self._clear_table_results()
        self._refresh_table_placeholder()
        self.status_text.set("表格队列已清空")

    def _refresh_table_placeholder(self) -> None:
        if self.table_list_placeholder is None:
            return
        if self.table_paths:
            self.table_list_placeholder.place_forget()
        else:
            self.table_list_placeholder.place(relx=0.5, rely=0.5, anchor="center")

    def choose_table_output(self) -> None:
        initial = Path(self.table_output_file.get()).expanduser() if self.table_output_file.get() else None
        initial_dir = initial.parent if initial else Path.cwd()
        initial_file = initial.name if initial else "表格汇总结果.xlsx"
        path = filedialog.asksaveasfilename(
            title="保存汇总结果",
            initialdir=str(initial_dir),
            initialfile=initial_file,
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if path:
            self.table_output_file.set(str(Path(path).resolve()))

    def toggle_table_advanced(self) -> None:
        visible = not self.table_advanced_visible.get()
        self.table_advanced_visible.set(visible)
        self.table_advanced_button_text.set("收起高级列匹配" if visible else "展开高级列匹配")
        if self.table_advanced_frame is not None:
            if visible:
                self.table_advanced_frame.grid()
            else:
                self.table_advanced_frame.grid_remove()

    def apply_selected_source_mapping(self) -> None:
        if self.table_selected_source_path is None:
            messagebox.showwarning("副表配置", "请先在副表列表中选择一张表。")
            return
        self.table_source_overrides[self.table_selected_source_path] = {
            "sheet": self.table_selected_source_sheet.get().strip(),
            "key": self.table_selected_source_key_column.get().strip(),
            "value": self.table_selected_source_value_column.get().strip(),
        }
        self._refresh_table_source_list_labels()
        self.status_text.set("已为选中副表保存特殊规则")

    def clear_selected_source_mapping(self) -> None:
        if self.table_selected_source_path is None:
            return
        self.table_source_overrides.pop(self.table_selected_source_path, None)
        self._load_selected_source_form(self.table_selected_source_path)
        self._refresh_table_source_list_labels()
        self.status_text.set("选中副表已改为使用默认规则")

    def remove_selected_table_setting(self) -> None:
        path = self._selected_table_setting_path()
        if path is None:
            self.remove_selected_tables()
            return
        if self.table_master_path is not None and path.resolve() == self.table_master_path.resolve():
            messagebox.showwarning("删除副表", "当前行是主表，请在左侧删除或先选择其他主表。")
            return
        if path in self.table_paths:
            index = self.table_paths.index(path)
            self.table_list.selection_clear(0, tk.END)
            self.table_list.selection_set(index)
            self.remove_selected_tables()

    def preview_table_merge(self) -> None:
        self.merge_tables(preview=True)

    def export_table_merge(self) -> None:
        self.merge_tables(preview=False)

    def inspect_tables(self) -> None:
        try:
            paths = self._table_inspection_paths()
            self._clear_table_results()
            self.status_text.set("正在检查表格...")

            def work():
                return TableWorkbookInspector().inspect_many(paths)

            def done(report):
                self._show_table_report(report)
                self.status_text.set(report.summary())

            self._run_background("表格检查", work, done)
        except Exception as exc:
            self.status_text.set("表格检查失败")
            messagebox.showerror("表格检查失败", str(exc))

    def merge_same_layout_tables(self) -> None:
        try:
            self._save_selected_table_settings(silent=True)
            master_path = self._table_master_path_from_form()
            source_paths = self._table_source_paths()
            output_path = self._table_output_path(master_path)
            self._clear_table_results()
            self.status_text.set("正在同格式汇总...")

            def work():
                return merge_same_layout(
                    master_path,
                    source_paths,
                    output_path,
                    master_sheet=self.table_master_sheet.get().strip() or None,
                    source_sheet=self.table_source_sheet.get().strip() or None,
                )

            def done(report):
                self._show_table_report(report)
                self.status_text.set(f"{report.summary()} 输出：{output_path}")

            self._run_background("同格式汇总", work, done)
        except Exception as exc:
            self.status_text.set("同格式汇总失败")
            messagebox.showerror("同格式汇总失败", str(exc))

    def merge_tables(self, *, preview: bool = False) -> None:
        try:
            self._save_selected_table_settings(silent=True)
            master_path = self._table_master_path_from_form()
            source_paths = self._table_source_paths()
            if preview:
                preview_dir = Path(tempfile.mkdtemp(prefix="office_tool_table_preview_"))
                output_path = preview_dir / f"{master_path.stem}_预览.xlsx"
            else:
                output_path = self._table_output_path(master_path)
            master_sheet = self.table_master_sheet.get().strip()
            master_key = self.table_master_key_column.get().strip()
            master_target = self.table_master_target_column.get().strip()
            source_sheet = self.table_source_sheet.get().strip()
            source_key = self.table_source_key_column.get().strip()
            source_value = self.table_source_value_column.get().strip()
            missing = [
                label
                for label, value in [
                    ("主表工作表", master_sheet),
                    ("主表校验列", master_key),
                    ("主表填入列", master_target),
                    ("副表校验列", source_key),
                    ("副表数据列", source_value),
                ]
                if not value
            ]
            if missing:
                raise ValueError("请先填写：" + "、".join(missing))
            self._clear_table_results()
            self.status_text.set("正在预览汇总..." if preview else "正在汇总表格...")

            def work():
                sources = [
                    self._source_mapping_for_path(path, source_sheet, source_key, source_value)
                    for path in source_paths
                ]
                return merge_by_columns(
                    TableMergeOptions(
                        master_path=master_path,
                        output_path=output_path,
                        master_sheet=master_sheet,
                        master_key_column=master_key,
                        master_target_column=master_target,
                        sources=sources,
                        fuzzy_match=self.table_fuzzy_enabled.get(),
                        fuzzy_threshold=self.table_fuzzy_threshold.get(),
                    )
                )

            def done(report):
                if preview:
                    self._preview_table_path(output_path, master_sheet or None)
                    self.status_text.set("汇总预览已生成")
                else:
                    self._show_table_report(report)
                    self.status_text.set(f"{report.summary()} 输出：{output_path}")

            self._run_background("汇总预览" if preview else "表格汇总", work, done)
        except Exception as exc:
            self.status_text.set("表格汇总失败")
            messagebox.showerror("表格汇总失败", str(exc))

    def _table_inspection_paths(self) -> list[Path]:
        paths: list[Path] = []
        master_path = self._table_master_path_from_form(required=False)
        if master_path is not None:
            paths.append(master_path)
        source_paths = self._table_source_paths(required=False)
        if source_paths:
            paths.extend(source_paths)
        if not paths:
            raise ValueError("请先选择主表，或添加副表文件。")
        return paths

    def _table_source_paths(self, *, required: bool = True) -> list[Path]:
        master = self._table_master_path_from_form(required=False)
        source_paths = collect_table_inputs(self.table_paths)
        if master is not None:
            source_paths = [path for path in source_paths if path.resolve() != master.resolve()]
        if required and not source_paths:
            raise ValueError("请先在左侧导入副表文件。")
        return source_paths

    def _table_master_path_from_form(self, required: bool = True) -> Path | None:
        if self.table_master_path is None:
            if required:
                raise ValueError("请先选择主表。")
            return None
        if not self.table_master_path.exists():
            raise FileNotFoundError(f"主表不存在: {self.table_master_path}")
        return self.table_master_path

    def _table_output_path(self, master_path: Path) -> Path:
        raw = self.table_output_file.get().strip()
        output = Path(raw).expanduser().resolve() if raw else master_path.with_name(f"{master_path.stem}_汇总结果.xlsx")
        self.table_output_file.set(str(output))
        return output

    def _load_table_workbook_info(self, path: Path) -> None:
        self.table_workbook_infos[path.resolve()] = inspect_workbook(path)

    def _refresh_master_table_choices(self) -> None:
        if self.table_master_path is None:
            return
        infos = self.table_workbook_infos.get(self.table_master_path.resolve()) or []
        sheet_names = [info.sheet for info in infos]
        if self.table_master_sheet_box is not None:
            self.table_master_sheet_box.configure(values=sheet_names)
        if sheet_names and self.table_master_sheet.get() not in sheet_names:
            self.table_master_sheet.set(sheet_names[0])
        self._refresh_master_header_choices()

    def _refresh_master_header_choices(self) -> None:
        headers = self._headers_for_path_sheet(self.table_master_path, self.table_master_sheet.get())
        for box in [self.table_master_key_box, self.table_master_target_box]:
            if box is not None:
                box.configure(values=headers)
        self._prefer_existing_or_first(self.table_master_key_column, headers, ["任务详情", "任务名称", "任务"])
        self._prefer_existing_or_first(self.table_master_target_column, headers, ["任务回复", "办理情况", "回复"])

    def _refresh_default_source_choices(self) -> None:
        sheet_names: list[str] = []
        headers: list[str] = []
        for path in self._table_source_paths(required=False):
            infos = self.table_workbook_infos.get(path.resolve()) or []
            for info in infos:
                if info.sheet not in sheet_names:
                    sheet_names.append(info.sheet)
                for header in info.headers:
                    if header not in headers:
                        headers.append(header)
        for box in [self.table_source_sheet_box]:
            if box is not None:
                box.configure(values=sheet_names)
        for box in [self.table_source_key_box, self.table_source_value_box]:
            if box is not None:
                box.configure(values=headers)
        if sheet_names and self.table_source_sheet.get() not in sheet_names:
            self.table_source_sheet.set(sheet_names[0])
        self._prefer_existing_or_first(self.table_source_key_column, headers, ["任务详情", "下达任务", "任务名称", "任务"])
        self._prefer_existing_or_first(self.table_source_value_column, headers, ["任务回复", "落实", "回复内容", "回复"])

    def _on_master_sheet_changed(self) -> None:
        self._refresh_master_header_choices()
        if self.table_master_path is not None:
            self._preview_table_path(self.table_master_path, self.table_master_sheet.get())

    def _on_default_source_sheet_changed(self) -> None:
        headers: list[str] = []
        for path in self.table_paths:
            headers.extend(header for header in self._headers_for_path_sheet(path, self.table_source_sheet.get()) if header not in headers)
        for box in [self.table_source_key_box, self.table_source_value_box]:
            if box is not None:
                box.configure(values=headers)
        self._prefer_existing_or_first(self.table_source_key_column, headers, ["任务详情", "下达任务", "任务名称", "任务"])
        self._prefer_existing_or_first(self.table_source_value_column, headers, ["任务回复", "落实", "回复内容", "回复"])

    def _on_selected_source_sheet_changed(self) -> None:
        if self.table_selected_source_path is None:
            return
        headers = self._headers_for_path_sheet(self.table_selected_source_path, self.table_selected_source_sheet.get())
        for box in [self.table_selected_source_key_box, self.table_selected_source_value_box]:
            if box is not None:
                box.configure(values=headers)
        self._prefer_existing_or_first(self.table_selected_source_key_column, headers, ["任务详情", "下达任务", "任务名称", "任务"])
        self._prefer_existing_or_first(self.table_selected_source_value_column, headers, ["任务回复", "落实", "回复内容", "回复"])
        self._preview_table_path(self.table_selected_source_path, self.table_selected_source_sheet.get())

    def _on_selected_table_setting_changed(self) -> None:
        path = self.table_selected_source_path
        if path is None:
            return
        headers = self._headers_for_path_sheet(path, self.table_selected_source_sheet.get())
        for box in [self.table_selected_source_key_box, self.table_selected_source_value_box]:
            if box is not None:
                box.configure(values=headers)
        self._prefer_existing_or_first(self.table_selected_source_key_column, headers, ["任务详情", "下达任务", "任务名称", "任务"])
        self._prefer_existing_or_first(self.table_selected_source_value_column, headers, ["任务回复", "落实", "回复内容", "回复"])
        self._save_selected_table_settings(silent=True)
        self._preview_table_path(path, self.table_selected_source_sheet.get())

    def _on_table_source_selected(self, _event=None) -> None:
        selection = self.table_list.curselection()
        if not selection:
            return
        path = self.table_paths[selection[0]]
        self.table_selected_source_path = path
        self._select_table_setting(path)
        self._load_selected_table_settings(path)
        self._preview_table_path(path, self.table_selected_source_sheet.get().strip() or None)

    def _on_table_settings_selected(self, _event=None) -> None:
        path = self._selected_table_setting_path()
        if path is None:
            return
        self.table_selected_source_path = path
        self._load_selected_table_settings(path)
        self._preview_table_path(path, self.table_selected_source_sheet.get().strip() or None)

    def _selected_table_setting_path(self) -> Path | None:
        if self.table_rules_tree is None:
            return None
        selection = self.table_rules_tree.selection()
        if not selection:
            return None
        return Path(selection[0])

    def _select_table_setting(self, path: Path) -> None:
        if self.table_rules_tree is None:
            return
        item = str(path.resolve())
        if self.table_rules_tree.exists(item):
            self.table_rules_tree.selection_set(item)
            self.table_rules_tree.see(item)

    def _load_selected_table_settings(self, path: Path) -> None:
        infos = self.table_workbook_infos.get(path.resolve()) or []
        sheets = [info.sheet for info in infos]
        is_master = self.table_master_path is not None and path.resolve() == self.table_master_path.resolve()
        if is_master:
            requested_sheet = self.table_master_sheet.get().strip()
            key = self.table_master_key_column.get()
            value = self.table_master_target_column.get()
        else:
            override = self.table_source_overrides.get(path, {})
            requested_sheet = override.get("sheet") or self.table_source_sheet.get().strip()
            key = override.get("key") or self.table_source_key_column.get()
            value = override.get("value") or self.table_source_value_column.get()
        sheet = requested_sheet if requested_sheet in sheets else (sheets[0] if sheets else "")
        headers = self._headers_for_path_sheet(path, sheet)
        self.table_selected_source_sheet.set(sheet)
        self.table_selected_source_key_column.set(key)
        self.table_selected_source_value_column.set(value)
        for box in [self.table_selected_source_sheet_box]:
            if box is not None:
                box.configure(values=sheets)
        for box in [self.table_selected_source_key_box, self.table_selected_source_value_box]:
            if box is not None:
                box.configure(values=headers)
        self._prefer_existing_or_first(self.table_selected_source_key_column, headers, ["任务详情", "下达任务", "任务名称", "任务"])
        self._prefer_existing_or_first(self.table_selected_source_value_column, headers, ["任务回复", "落实", "回复内容", "办理情况", "回复"])

    def _save_selected_table_settings(self, *, silent: bool = False) -> None:
        path = self.table_selected_source_path
        if path is None:
            return
        is_master = self.table_master_path is not None and path.resolve() == self.table_master_path.resolve()
        if is_master:
            self.table_master_sheet.set(self.table_selected_source_sheet.get().strip())
            self.table_master_key_column.set(self.table_selected_source_key_column.get().strip())
            self.table_master_target_column.set(self.table_selected_source_value_column.get().strip())
        else:
            self.table_source_overrides[path] = {
                "sheet": self.table_selected_source_sheet.get().strip(),
                "key": self.table_selected_source_key_column.get().strip(),
                "value": self.table_selected_source_value_column.get().strip(),
            }
        self._refresh_table_source_list_labels()
        self._select_table_setting(path)
        if not silent:
            self.status_text.set("设置已更新")

    def _load_selected_source_form(self, path: Path) -> None:
        override = self.table_source_overrides.get(path, {})
        infos = self.table_workbook_infos.get(path.resolve()) or []
        sheets = [info.sheet for info in infos]
        requested_sheet = override.get("sheet") or self.table_source_sheet.get().strip()
        sheet = requested_sheet if requested_sheet in sheets else (sheets[0] if sheets else "")
        self.table_selected_source_sheet.set(sheet)
        headers = self._headers_for_path_sheet(path, sheet)
        self.table_selected_source_key_column.set(override.get("key") or self.table_source_key_column.get())
        self.table_selected_source_value_column.set(override.get("value") or self.table_source_value_column.get())
        if self.table_selected_source_sheet_box is not None:
            self.table_selected_source_sheet_box.configure(values=sheets)
        for box in [self.table_selected_source_key_box, self.table_selected_source_value_box]:
            if box is not None:
                box.configure(values=headers)

    def _clear_selected_source_form(self) -> None:
        self.table_selected_source_sheet.set("")
        self.table_selected_source_key_column.set("")
        self.table_selected_source_value_column.set("")
        for box in [self.table_selected_source_sheet_box, self.table_selected_source_key_box, self.table_selected_source_value_box]:
            if box is not None:
                box.configure(values=[])

    def _source_mapping_for_path(self, path: Path, default_sheet: str, default_key: str, default_value: str) -> SourceColumnMapping:
        override = self.table_source_overrides.get(path, {})
        sheet = override.get("sheet") or default_sheet or _first_sheet_name(path)
        key = override.get("key") or default_key
        value = override.get("value") or default_value
        if not key or not value:
            raise ValueError(f"请为副表配置校验列和数据列: {path.name}")
        return SourceColumnMapping(path=path, sheet=sheet, key_column=key, value_column=value)

    def _headers_for_path_sheet(self, path: Path | None, sheet_name: str) -> list[str]:
        if path is None or not sheet_name:
            return []
        for info in self.table_workbook_infos.get(path.resolve(), []):
            if info.sheet == sheet_name:
                return list(info.headers)
        return []

    @staticmethod
    def _prefer_existing_or_first(variable: tk.StringVar, choices: list[str], preferred: list[str]) -> None:
        if not choices:
            return
        if variable.get() in choices:
            return
        for item in preferred:
            if item in choices:
                variable.set(item)
                return
        variable.set(choices[0])

    def _refresh_table_source_list_labels(self) -> None:
        self.table_list.delete(0, tk.END)
        master = self.table_master_path.resolve() if self.table_master_path is not None else None
        for path in self.table_paths:
            markers: list[str] = []
            if master is not None and path.resolve() == master:
                markers.append("主表")
            if path in self.table_source_overrides:
                markers.append("特殊")
            prefix = f"[{'/'.join(markers)}] " if markers else ""
            self.table_list.insert(tk.END, prefix + path.name)
        self._refresh_table_rules_tree()

    def _refresh_table_rules_tree(self) -> None:
        if self.table_rules_tree is None:
            return
        selection = self.table_rules_tree.selection()
        selected = selection[0] if selection else ""
        for item in self.table_rules_tree.get_children():
            self.table_rules_tree.delete(item)
        master = self.table_master_path.resolve() if self.table_master_path is not None else None
        for path in self.table_paths:
            resolved = path.resolve()
            is_master = master is not None and resolved == master
            if is_master:
                sheet = self.table_master_sheet.get()
                key = self.table_master_key_column.get()
                value = self.table_master_target_column.get()
            else:
                override = self.table_source_overrides.get(path, {})
                sheet = override.get("sheet") or self.table_source_sheet.get()
                key = override.get("key") or self.table_source_key_column.get()
                value = override.get("value") or self.table_source_value_column.get()
            item_id = str(resolved)
            self.table_rules_tree.insert(
                "",
                tk.END,
                iid=item_id,
                values=(
                    "主表" if is_master else "副表",
                    path.name,
                    sheet,
                    key,
                    value,
                ),
            )
        if selected and self.table_rules_tree.exists(selected):
            self.table_rules_tree.selection_set(selected)

    def add_document_folder(self) -> None:
        folder = filedialog.askdirectory(title="选择文件夹")
        if folder:
            self._add_document_paths([folder])

    def _add_document_paths(self, paths) -> None:
        for raw in paths:
            path = Path(raw).resolve()
            if path not in self.document_paths:
                self.document_paths.append(path)
                self.doc_list.insert(tk.END, path.name)
        if paths:
            self._refresh_default_document_targets()
        self._refresh_document_placeholder()
        self.status_text.set(f"队列中 {len(self.document_paths)} 个路径")

    def remove_selected_documents(self) -> None:
        for index in reversed(list(self.doc_list.curselection())):
            self.doc_list.delete(index)
            del self.document_paths[index]
        self._refresh_default_document_targets()
        self._refresh_document_placeholder()
        self.status_text.set(f"队列中 {len(self.document_paths)} 个路径")

    def clear_documents(self) -> None:
        self.doc_list.delete(0, tk.END)
        self.document_paths.clear()
        self.doc_output_dir.set("")
        self.doc_report_dir.set("")
        self._refresh_document_placeholder()
        self.status_text.set("队列已清空")

    def _refresh_document_placeholder(self) -> None:
        if self.doc_list_placeholder is None:
            return
        if self.document_paths:
            self.doc_list_placeholder.place_forget()
        else:
            self.doc_list_placeholder.place(relx=0.5, rely=0.5, anchor="center")

    def _default_output_root(self, paths: list[Path]) -> Path:
        first = paths[0].resolve()
        base = first if first.is_dir() else first.parent
        if self._is_single_document_source(paths):
            return base / f"{first.stem}_formatted.docx"
        return base / "formatted_output"

    @staticmethod
    def _is_single_document_source(paths: list[Path]) -> bool:
        return len(paths) == 1 and (paths[0].is_file() or paths[0].suffix.lower() in {".doc", ".docx", ".txt", ".md"})

    def _refresh_default_document_targets(self) -> None:
        if not self.document_paths:
            return
        default_output = self._default_output_root(self.document_paths)
        self.doc_output_dir.set(str(default_output))
        first = self.document_paths[0].resolve()
        base = first if first.is_dir() else first.parent
        self.doc_report_dir.set(str(base / "proofreading_reports"))

    def _ask_format_output_target(self, paths: list[Path]) -> Path | None:
        default = self._default_output_root(paths)
        current = Path(self.doc_output_dir.get()).expanduser() if self.doc_output_dir.get() else None
        if current and not (self._is_single_document_source(paths) and current.suffix.lower() == ".docx"):
            default = current
        if self._is_single_document_source(paths):
            initial_dir = default.parent if default.suffix else default
            initial_file = f"{paths[0].stem}_formatted.docx"
            path = filedialog.asksaveasfilename(
                title="保存校对后的文件",
                initialdir=str(initial_dir),
                initialfile=initial_file,
                defaultextension=".docx",
                filetypes=[("Word 文档", "*.docx")],
            )
            return Path(path).resolve() if path else None
        folder = filedialog.askdirectory(title="选择批量导出保存目录", initialdir=str(default if default.suffix == "" else default.parent))
        return Path(folder).resolve() if folder else None

    def _ensure_report_dir(self, paths: list[Path]) -> None:
        if self.doc_report_dir.get():
            return
        first = paths[0].resolve()
        base = first if first.is_dir() else first.parent
        self.doc_report_dir.set(str(base / "proofreading_reports"))

    def _load_ai_profiles(self) -> None:
        self.ai_profiles = {"默认配置": AIReviewOptions()}
        active = "默认配置"
        needs_migration = False
        if AI_PROFILE_FILE.exists():
            try:
                raw = json.loads(AI_PROFILE_FILE.read_text(encoding="utf-8"))
                active = raw.get("active") or active
                profiles = raw.get("profiles") or {}
                for name, value in profiles.items():
                    if isinstance(value, dict) and name:
                        allowed = AIReviewOptions().__dict__
                        merged = {key: value.get(key, default) for key, default in allowed.items()}
                        protected_key = value.get("api_key_protected", "")
                        if protected_key:
                            merged["api_key"] = unprotect_secret(str(protected_key))
                        elif merged.get("api_key"):
                            needs_migration = True
                        self.ai_profiles[name] = AIReviewOptions(**merged)
            except Exception:
                self.ai_profiles = {"默认配置": AIReviewOptions()}
                active = "默认配置"
        if active not in self.ai_profiles:
            active = next(iter(self.ai_profiles))
        self.ai_profile_name.set(active)
        self.config.ai_review = self.ai_profiles[active]
        self._sync_form_from_config()
        if needs_migration:
            self._save_ai_profiles()

    def _save_ai_profiles(self) -> None:
        AI_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
        profiles = {}
        for name, options in self.ai_profiles.items():
            value = dict(options.__dict__)
            api_key = value.pop("api_key", "")
            value["api_key_protected"] = protect_secret(api_key)
            profiles[name] = value
        raw = {
            "active": self.ai_profile_name.get().strip() or "默认配置",
            "profiles": profiles,
        }
        AI_PROFILE_FILE.write_text(json.dumps(raw, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    def _ai_options_from_form(self) -> AIReviewOptions:
        return AIReviewOptions(
            enabled=self.ai_enabled.get(),
            base_url=self.ai_base_url.get().strip(),
            model=self.ai_model.get().strip() or "deepseek-chat",
            api_key=self.ai_api_key.get().strip(),
            api_key_env=self.ai_key_env.get().strip() or "DEEPSEEK_API_KEY",
            auth_prefix=self.ai_auth_prefix.get().strip(),
            stream=self.ai_stream.get(),
        )

    def _apply_ai_options_to_form(self, options: AIReviewOptions) -> None:
        self.ai_enabled.set(options.enabled)
        self.ai_base_url.set(options.base_url)
        self.ai_model.set(options.model)
        self.ai_api_key.set(options.api_key)
        self.ai_key_env.set(options.api_key_env)
        self.ai_auth_prefix.set(options.auth_prefix)
        self.ai_stream.set(options.stream)
        self.config.ai_review = options
        self._refresh_ai_toggle_text()

    def _refresh_ai_toggle_text(self) -> None:
        self.ai_toggle_text.set("关闭 AI" if self.ai_enabled.get() else "启用 AI")

    def toggle_ai_review(self) -> None:
        self.ai_enabled.set(not self.ai_enabled.get())
        self.config.ai_review.enabled = self.ai_enabled.get()
        self._refresh_ai_toggle_text()
        self.status_text.set("AI 校对已启用" if self.ai_enabled.get() else "AI 校对已关闭")

    def open_ai_settings(self) -> None:
        dialog = tk.Toplevel(self.root)
        dialog.title("AI 配置")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=COLOR["surface"])
        dialog.geometry("820x720")
        dialog.minsize(760, 680)
        body = ttk.Frame(dialog, style="Card.TFrame", padding=(22, 18, 22, 18))
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(1, weight=1)

        ttk.Label(body, text="AI 文本校对", style="CardTitle.TLabel").grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 14))
        ttk.Label(body, text="配置档", style="Card.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        profile_combo = ttk.Combobox(body, textvariable=self.ai_profile_name, values=sorted(self.ai_profiles), state="normal")
        profile_combo.grid(row=1, column=1, columnspan=2, sticky="ew", pady=6)

        self._dialog_field(body, 2, "服务地址", self.ai_base_url, "http://deepseek.local:8000/v1 或完整 /chat/completions 地址")
        self._dialog_field(body, 3, "模型名称", self.ai_model, "例如 deepseek-chat、qwen、glm 等")
        self._dialog_field(body, 4, "API Key", self.ai_api_key, "留空则从环境变量读取", show="*")
        self._dialog_field(body, 5, "环境变量名", self.ai_key_env, "默认 DEEPSEEK_API_KEY")
        self._dialog_field(body, 6, "鉴权前缀", self.ai_auth_prefix, "默认 Bearer；留空则直接发送 token")
        ttk.Checkbutton(body, text="使用流式响应解析", variable=self.ai_stream).grid(row=7, column=1, columnspan=2, sticky="w", pady=(6, 12))

        buttons = ttk.Frame(body, style="Card.TFrame")
        buttons.grid(row=8, column=0, columnspan=3, sticky="ew", pady=(10, 0))
        for column in range(4):
            buttons.columnconfigure(column, weight=1, uniform="ai_buttons")

        def refresh_profiles() -> None:
            profile_combo.configure(values=sorted(self.ai_profiles))

        def select_profile(_event=None) -> None:
            name = self.ai_profile_name.get().strip()
            if name in self.ai_profiles:
                self._apply_ai_options_to_form(self.ai_profiles[name])

        def save_profile() -> None:
            name = self.ai_profile_name.get().strip() or "默认配置"
            options = self._ai_options_from_form()
            self.ai_profile_name.set(name)
            self.ai_profiles[name] = options
            self.config.ai_review = options
            self._save_ai_profiles()
            refresh_profiles()
            self.status_text.set("AI 配置已保存")
            messagebox.showinfo("AI 配置", f"已保存配置档：{name}", parent=dialog)

        def delete_profile() -> None:
            name = self.ai_profile_name.get().strip()
            if name == "默认配置":
                messagebox.showinfo("AI 配置", "默认配置不能删除。", parent=dialog)
                return
            if name in self.ai_profiles:
                del self.ai_profiles[name]
                self.ai_profile_name.set("默认配置")
                self._apply_ai_options_to_form(self.ai_profiles["默认配置"])
                self._save_ai_profiles()
                refresh_profiles()
                self.status_text.set("AI 配置已删除")

        profile_combo.bind("<<ComboboxSelected>>", select_profile)
        ttk.Button(buttons, text="保存配置", style="Primary.TButton", command=save_profile).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ttk.Button(buttons, text="测试 AI 连接", style="Secondary.TButton", command=self.test_ai_connection).grid(row=0, column=1, sticky="ew", padx=(0, 8))
        ttk.Button(buttons, text="删除配置", style="Subtle.TButton", command=delete_profile).grid(row=0, column=2, sticky="ew", padx=(0, 8))
        ttk.Button(buttons, text="关闭", style="Subtle.TButton", command=dialog.destroy).grid(row=0, column=3, sticky="ew")

    def _dialog_field(self, parent, row, label, var, hint: str = "", show: str | None = None) -> None:
        ttk.Label(parent, text=label, style="Card.TLabel").grid(row=row, column=0, sticky="w", padx=(0, 10), pady=5)
        field = ttk.Frame(parent, style="Card.TFrame")
        field.grid(row=row, column=1, columnspan=2, sticky="ew", pady=5)
        field.columnconfigure(0, weight=1)
        ttk.Entry(field, textvariable=var, show=show or "").grid(row=0, column=0, sticky="ew")
        if hint:
            ttk.Label(field, text=hint, style="Muted.TLabel").grid(row=1, column=0, sticky="w", pady=(3, 0))

    def load_config(self) -> None:
        path = filedialog.askopenfilename(title="选择配置文件", filetypes=[("JSON 配置", "*.json"), ("所有文件", "*.*")])
        if not path:
            return
        current_key = self.config.ai_review.api_key
        self.config = OfficeToolConfig.from_dict(json.loads(Path(path).read_text(encoding="utf-8")))
        if not self.config.ai_review.api_key:
            self.config.ai_review.api_key = current_key
        self._sync_form_from_config()
        self.status_text.set("配置已载入")

    def save_config(self) -> None:
        path = filedialog.asksaveasfilename(title="导出配置", defaultextension=".json", filetypes=[("JSON 配置", "*.json")])
        if not path:
            return
        self._sync_config_from_form()
        self._write_external_config(path)
        name = Path(path).stem.strip() or "自定义配置"
        if name in BUILTIN_SCHEMES:
            name += "（自定义）"
        if name in self.config_profiles and not messagebox.askyesno("覆盖配置", f"程序内已存在“{name}”，是否同步覆盖？"):
            self.status_text.set("配置已导出，程序内副本未覆盖")
            return
        self.config_profile_store.save(name, self.config)
        self.config_profiles = self.config_profile_store.load_all()
        self._refresh_config_profile_values()
        self.status_text.set(f"配置已导出，并保存程序内副本“{name}”")

    def rename_config_profile(self, old_name: str) -> None:
        new_name = simpledialog.askstring("重命名配置", "新名称", initialvalue=old_name, parent=self.root)
        if new_name is None or new_name.strip() == old_name:
            return
        new_name = new_name.strip()
        if not new_name or new_name in BUILTIN_SCHEMES:
            messagebox.showwarning("重命名配置", "请输入未被默认方案占用的名称。")
            return
        if new_name in self.config_profiles and not messagebox.askyesno("覆盖配置", f"“{new_name}”已存在，是否覆盖？"):
            return
        self.config_profile_store.rename(old_name, new_name)
        self.config_profiles = self.config_profile_store.load_all()
        self._refresh_config_profile_values()
        self.status_text.set(f"配置已重命名为“{new_name}”")

    def delete_config_profile(self, name: str) -> None:
        if not messagebox.askyesno("删除配置", f"确定删除“{name}”吗？"):
            return
        self.config_profile_store.delete(name)
        self.config_profiles = self.config_profile_store.load_all()
        self._refresh_config_profile_values()
        self.status_text.set("自定义配置已删除")

    def _refresh_config_profile_values(self) -> None:
        if self.config_profile_popup is not None and self.config_profile_popup.winfo_exists():
            self._render_config_profile_popup(self.config_profile_popup)
            self._position_config_profile_popup()

    def open_config_profile_popup(self, anchor: tk.Widget) -> None:
        if self.config_profile_popup is not None and self.config_profile_popup.winfo_exists():
            self._close_config_profile_popup()
            return
        popup = tk.Toplevel(self.root)
        popup.withdraw()
        popup.wm_overrideredirect(True)
        popup.transient(self.root)
        popup.configure(bg=COLOR["border"])
        self.config_profile_popup = popup
        self.config_profile_anchor = anchor
        self._render_config_profile_popup(popup)
        popup.update_idletasks()
        self._position_config_profile_popup()
        popup.deiconify()
        popup.lift()
        self.root.after_idle(self._position_config_profile_popup)
        self.config_popup_bind_id = self.root.bind(
            "<Configure>",
            lambda _event: self.root.after_idle(self._position_config_profile_popup),
            add="+",
        )
        for sequence in ("<ButtonPress-1>", "<ButtonPress-2>", "<ButtonPress-3>"):
            bind_id = self.root.bind(sequence, self._handle_config_popup_root_click, add="+")
            if bind_id is not None:
                self.config_popup_click_bind_ids[sequence] = bind_id
        popup.bind("<Escape>", lambda _event: self._close_config_profile_popup())
        popup.bind("<FocusOut>", self._handle_config_popup_focus_out, add="+")
        popup.focus_force()

    def _handle_config_popup_root_click(self, event) -> None:
        if _widget_is_within(event.widget, self.config_profile_anchor):
            return
        self._close_config_profile_popup()

    def _handle_config_popup_focus_out(self, _event=None) -> None:
        self.root.after_idle(self._close_config_popup_if_focus_outside)

    def _close_config_popup_if_focus_outside(self) -> None:
        popup = self.config_profile_popup
        if popup is None or not popup.winfo_exists():
            return
        focused = self.root.focus_get()
        if _widget_is_within(focused, popup) or _widget_is_within(focused, self.config_profile_anchor):
            return
        self._close_config_profile_popup()

    def _position_config_profile_popup(self) -> None:
        popup = self.config_profile_popup
        anchor = self.config_profile_anchor
        if popup is None or anchor is None or not popup.winfo_exists() or not anchor.winfo_exists():
            return
        anchor.update_idletasks()
        popup.update_idletasks()
        width, height, x, y = _attached_popup_geometry(
            anchor_x=anchor.winfo_rootx(),
            anchor_y=anchor.winfo_rooty(),
            anchor_width=anchor.winfo_width(),
            anchor_height=anchor.winfo_height(),
            requested_height=popup.winfo_reqheight(),
            screen_x=popup.winfo_vrootx(),
            screen_y=popup.winfo_vrooty(),
            screen_width=popup.winfo_vrootwidth(),
            screen_height=popup.winfo_vrootheight(),
        )
        popup.wm_geometry(f"{width}x{height}{x:+d}{y:+d}")
        popup.lift()

    def _close_config_profile_popup(self) -> None:
        popup = self.config_profile_popup
        self.config_profile_popup = None
        self.config_profile_anchor = None
        if self.config_popup_bind_id is not None:
            self.root.unbind("<Configure>", self.config_popup_bind_id)
            self.config_popup_bind_id = None
        for sequence, bind_id in self.config_popup_click_bind_ids.items():
            self.root.unbind(sequence, bind_id)
        self.config_popup_click_bind_ids.clear()
        if popup is not None and popup.winfo_exists():
            popup.destroy()

    def _run_config_menu_action(self, action) -> None:
        self._close_config_profile_popup()
        action()

    def _render_config_profile_popup(self, popup: tk.Toplevel) -> None:
        for child in popup.winfo_children():
            child.destroy()
        panel = ttk.Frame(popup, style="Card.TFrame", padding=(5, 5, 5, 5))
        panel.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        panel.columnconfigure(0, weight=1)
        ttk.Button(panel, text="导入配置", style="Dropdown.TButton", command=lambda: self._run_config_menu_action(self.load_config)).grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 2))
        ttk.Button(panel, text="导出配置", style="Dropdown.TButton", command=lambda: self._run_config_menu_action(self.save_config)).grid(row=1, column=0, columnspan=3, sticky="ew", pady=(2, 4))
        ttk.Separator(panel, orient=tk.HORIZONTAL).grid(row=2, column=0, columnspan=3, sticky="ew", pady=(2, 5))
        if not self.config_profiles:
            ttk.Label(panel, text="暂无已导出的自定义配置", style="Muted.TLabel").grid(row=3, column=0, columnspan=3, sticky="w", padx=5, pady=6)
            return
        for index, name in enumerate(sorted(self.config_profiles), start=3):
            apply_button = ttk.Button(panel, text=name, style="Dropdown.TButton", command=lambda value=name: self._apply_custom_profile(value))
            apply_button.grid(row=index, column=0, sticky="ew", pady=2)
            rename_button = ttk.Button(panel, text="✎", width=3, style="Quiet.TButton", command=lambda value=name: self.rename_config_profile(value))
            rename_button.grid(row=index, column=1, padx=(6, 2), pady=2)
            delete_button = ttk.Button(panel, text="×", width=3, style="Quiet.TButton", command=lambda value=name: self.delete_config_profile(value))
            delete_button.grid(row=index, column=2, padx=(2, 0), pady=2)
            _attach_tooltip(rename_button, "重命名程序内副本")
            _attach_tooltip(delete_button, "删除程序内副本")

    def _apply_custom_profile(self, name: str) -> None:
        self._close_config_profile_popup()
        self._apply_scheme_callback(name)

    def _write_external_config(self, path: str | Path) -> None:
        raw = self.config.to_dict()
        raw["ai_review"]["api_key"] = ""
        Path(path).write_text(json.dumps(raw, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    def init_config(self) -> None:
        path = filedialog.asksaveasfilename(title="生成默认配置", defaultextension=".json", filetypes=[("JSON 配置", "*.json")])
        if not path:
            return
        Path(path).write_text(json.dumps(OfficeToolConfig().to_dict(), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        self.status_text.set("默认配置已生成")

    def audit_documents(self) -> None:
        try:
            self._sync_config_from_form()
            paths = self._require_documents()
            self._ensure_report_dir(paths)
            self._clear_results()
            config = deepcopy(self.config)
            report_dir = self.doc_report_dir.get() or None
            self.status_text.set("正在校对...")
            if config.ai_review.enabled:
                self.status_text.set("正在校对（AI 已启用）...")

            def work():
                return audit_many(paths, config, report_dir, markdown=True)

            def done(results):
                self._show_document_results(results)
                self.status_text.set(summarize_results(results))

            self._run_background("校对", work, done)
        except Exception as exc:
            self.status_text.set("校对失败")
            messagebox.showerror("校对失败", str(exc))

    def format_documents(self) -> None:
        try:
            self._sync_config_from_form()
            paths = list(self.document_paths)
            editor_text = _text_without_placeholder(self.direct_text)
            editor_source: Path | None = None
            if editor_text:
                documents_dir = Path.home() / "Documents"
                editor_source = (documents_dir if documents_dir.exists() else Path.home()) / "直接输入文稿.txt"
            display_paths = [*paths, *([editor_source] if editor_source is not None else [])]
            if not display_paths:
                raise ValueError("请先添加文件，或在文本输入框中填写内容。")
            self._validate_generation_config()
            if (
                len(display_paths) > 1
                and self.config.generation.add_red_head
                and self.config.audit.profile in {"red_head", "letter_head"}
                and self.config.generation.document_number
                and not messagebox.askyesno(
                    "批量添加红头",
                    f"将为 {len(display_paths)} 份任务使用同一发文字号“{self.config.generation.document_number}”，是否继续？",
                )
            ):
                self.status_text.set("已取消导出")
                return
            output = self._ask_format_output_target(display_paths)
            if not output:
                self.status_text.set("已取消导出")
                return
            self.doc_output_dir.set(str(output))
            self._ensure_report_dir(display_paths)
            self._clear_results()
            config = deepcopy(self.config)
            report_dir = self.doc_report_dir.get() or None
            self.status_text.set("正在校对导出...")
            if config.ai_review.enabled:
                self.status_text.set("正在校对导出（AI 已启用）...")

            def work():
                if not editor_text:
                    return format_many(paths, output, config, report_dir=report_dir, markdown=True)
                with tempfile.TemporaryDirectory(prefix="office_tool_direct_text_") as tmp:
                    temporary_source = Path(tmp) / "直接输入文稿.txt"
                    temporary_source.write_text(editor_text, encoding="utf-8")
                    results = format_many(
                        [*paths, temporary_source],
                        output,
                        config,
                        report_dir=report_dir,
                        markdown=True,
                    )
                    for result in results:
                        if result.source == temporary_source.resolve() and editor_source is not None:
                            result.source = editor_source
                    return results

            def done(results):
                self._show_document_results(results)
                self.status_text.set(summarize_results(results))

            self._run_background("校对导出", work, done)
        except Exception as exc:
            self.status_text.set("校对导出失败")
            messagebox.showerror("校对导出失败", str(exc))

    def _validate_generation_config(self) -> None:
        options = self.config.generation
        profile = self.config.audit.profile
        missing: list[str] = []
        if options.add_red_head and profile in {"red_head", "letter_head"}:
            if not options.red_head_title:
                missing.append("红头名称")
        if options.add_red_head and profile == "meeting_minutes":
            for label, value in [
                ("会议期号", options.meeting_number),
                ("编发单位", options.meeting_organization),
                ("编发日期", options.meeting_date),
            ]:
                if not value:
                    missing.append(label)
        if options.add_imprint and profile == "red_head":
            for label, value in [
                ("印发单位", options.print_organization),
                ("印发日期", options.print_date),
            ]:
                if not value:
                    missing.append(label)
        if options.add_imprint and profile == "meeting_minutes" and not options.distribution:
            missing.append("分送内容")
        if missing:
            raise ValueError("添加内容前请填写：" + "、".join(missing))

    def test_ai_connection(self) -> None:
        try:
            self._sync_config_from_form()
            if not self.config.ai_review.base_url:
                raise ValueError("请先填写 DeepSeek 服务地址。")
            options = deepcopy(self.config.ai_review)
            self.status_text.set("正在测试 AI 连接...")

            def work():
                reviewer = DeepSeekTextReviewer(options)
                return reviewer.review_text("测试标题\n主送机关：示例单位\n这是一段用于验证连通性的测试文本。")

            def done(findings):
                self.status_text.set("AI 连接成功")
                messagebox.showinfo("AI 连接", f"DeepSeek 连接成功，解析到 {len(findings)} 条建议。")

            self._run_background("AI 连接", work, done)
        except Exception as exc:
            self.status_text.set("AI 连接失败")
            messagebox.showerror("AI 连接失败", str(exc))

    def _run_background(self, label: str, work, done) -> None:
        def runner() -> None:
            try:
                result = work()
            except Exception as exc:
                self.root.after(0, lambda error=exc: self._background_failed(label, error))
                return
            self.root.after(0, lambda: done(result))

        threading.Thread(target=runner, daemon=True).start()

    def _background_failed(self, label: str, exc: Exception) -> None:
        self.status_text.set(f"{label}失败")
        messagebox.showerror(f"{label}失败", str(exc))

    def _sync_form_from_config(self) -> None:
        self.scheme_var.set(_profile_label(self.config.audit.profile))
        self.ai_enabled.set(self.config.ai_review.enabled)
        self.ai_base_url.set(self.config.ai_review.base_url)
        self.ai_model.set(self.config.ai_review.model)
        self.ai_api_key.set(self.config.ai_review.api_key)
        self.ai_key_env.set(self.config.ai_review.api_key_env)
        self.ai_auth_prefix.set(self.config.ai_review.auth_prefix)
        self.ai_stream.set(self.config.ai_review.stream)
        self._refresh_ai_toggle_text()
        for key, var in self.page_vars.items():
            var.set(str(getattr(self.config.page, key)))
        for key, var in self.format_vars.items():
            var.set(bool(getattr(self.config.format, key)))
        for key, var in self.audit_vars.items():
            var.set(bool(getattr(self.config.audit, key)))
        for key, (font_var, size_var) in self.style_vars.items():
            style = self.config.styles[key]
            font_var.set(style.font)
            size_var.set(_font_size_label(style.size_pt))
        for key, var in self.generation_vars.items():
            var.set(getattr(self.config.generation, key))
        self._refresh_generation_panel()

    def _sync_config_from_form(self) -> None:
        config = self.config
        config.ai_review.enabled = self.ai_enabled.get()
        config.ai_review.base_url = self.ai_base_url.get().strip()
        config.ai_review.model = self.ai_model.get().strip() or "deepseek-chat"
        config.ai_review.api_key = self.ai_api_key.get().strip()
        config.ai_review.api_key_env = self.ai_key_env.get().strip() or "DEEPSEEK_API_KEY"
        config.ai_review.auth_prefix = self.ai_auth_prefix.get().strip()
        config.ai_review.stream = self.ai_stream.get()
        for key, var in self.page_vars.items():
            current = getattr(config.page, key)
            raw = var.get().strip()
            setattr(config.page, key, int(raw) if isinstance(current, int) else float(raw))
        for key, var in self.format_vars.items():
            setattr(config.format, key, var.get())
        for key, var in self.audit_vars.items():
            setattr(config.audit, key, var.get())
        for key, (font_var, size_var) in self.style_vars.items():
            config.styles[key].font = font_var.get().strip()
            config.styles[key].size_pt = _font_size_value(size_var.get())
        for key, var in self.generation_vars.items():
            value = var.get()
            setattr(config.generation, key, value.strip() if isinstance(value, str) else bool(value))

    def _require_documents(self) -> list[Path]:
        if not self.document_paths:
            raise ValueError("请先添加文件或文件夹。")
        return list(self.document_paths)

    def _show_document_results(self, results) -> None:
        for result in results:
            if result.error:
                item = self.result_tree.insert("", tk.END, values=(result.source.name, "运行错误", "运行", self._shorten(result.error)))
                self.result_details[item] = f"文件：{result.source}\n错误类型：运行错误\n来源：运行\n\n完整内容：\n{result.error}"
                continue
            if result.report:
                self._insert_report_findings(result.source.name, result.report)
        self._refresh_result_placeholder()

    def _insert_report_findings(self, file_name: str, report) -> None:
        if not report.findings:
            item = self.result_tree.insert("", tk.END, values=(file_name, "通过", "规则", "未发现问题"))
            self.result_details[item] = f"文件：{file_name}\n错误类型：通过\n来源：规则\n\n未发现问题。"
            return
        for finding in report.findings:
            source = "AI 建议" if finding.code.startswith("ai_") else "确定性规则"
            finding_type = self._finding_type(finding)
            message = finding.message
            if finding.block_index is not None:
                message = f"第 {finding.block_index + 1} 段：{message}"
            item = self.result_tree.insert(
                "",
                tk.END,
                values=(file_name, finding_type, source, self._shorten(message)),
            )
            self.result_details[item] = self._format_finding_detail(file_name, finding_type, source, finding)

    def _finding_type(self, finding) -> str:
        if finding.code.startswith("layout_"):
            return "格式错误"
        if finding.code.startswith("ai_"):
            raw = finding.code[3:].replace("_", " ")
            text = raw + " " + finding.message + " " + finding.suggestion
            for keyword, finding_type in AI_TYPE_KEYWORDS.items():
                if keyword in text:
                    return finding_type
            return raw.strip() or "AI 建议"
        return FINDING_TYPE_BY_CODE.get(finding.code, "校对提示")

    def _format_finding_detail(self, file_name: str, finding_type: str, source: str, finding) -> str:
        parts = [
            f"文件：{file_name}",
            f"错误类型：{finding_type}",
            f"来源：{source}",
        ]
        if finding.block_index is not None:
            parts.append(f"段落：第 {finding.block_index + 1} 段")
        parts.append("")
        parts.append("问题说明：")
        parts.append(finding.message or "无")
        if finding.text:
            parts.extend(["", "原文引用：", finding.text])
        if finding.expected:
            parts.extend(["", "期望：", finding.expected])
        if finding.actual:
            parts.extend(["", "实际：", finding.actual])
        if finding.suggestion:
            parts.extend(["", "处理建议：", finding.suggestion])
        return "\n".join(parts)

    def _show_selected_result(self, _event=None) -> None:
        selection = self.result_tree.selection()
        if not selection:
            return
        self._set_result_detail(self.result_details.get(selection[0], ""))

    def _set_result_detail(self, text: str) -> None:
        self.result_detail_text.configure(state="normal")
        self.result_detail_text.delete("1.0", tk.END)
        self.result_detail_text.insert(tk.END, text)
        self.result_detail_text.configure(state="disabled")

    @staticmethod
    def _shorten(text: str, limit: int = 72) -> str:
        normalized = " ".join(str(text).split())
        if len(normalized) <= limit:
            return normalized
        return normalized[: limit - 1] + "…"

    def _clear_results(self) -> None:
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        self.result_details.clear()
        self._set_result_detail("选中上方任意一条校对结果，可查看完整内容。")
        self._refresh_result_placeholder()

    def _refresh_result_placeholder(self) -> None:
        if self.result_tree_placeholder is None:
            return
        if self.result_tree.get_children():
            self.result_tree_placeholder.place_forget()
        else:
            self.result_tree_placeholder.place(relx=0.5, rely=0.5, anchor="center")

    def _show_table_report(self, report) -> None:
        rows: list[list[str]] = []
        for sheet in report.sheets:
            merged = "、".join(sheet.merged_ranges[:3])
            if len(sheet.merged_ranges) > 3:
                merged += f" 等 {len(sheet.merged_ranges)} 项"
            rows.append(
                [
                    "结构",
                    Path(sheet.workbook).name,
                    sheet.sheet,
                    f"表头第 {sheet.header_row} 行",
                    f"{len(sheet.headers)} 列；合并单元格：{merged or '无'}",
                ]
            )

        for finding in report.findings:
            rows.append(
                [
                    finding.severity,
                    Path(finding.workbook).name if finding.workbook else "",
                    finding.sheet,
                    self._table_finding_location(finding),
                    finding.message if not finding.actual else f"{finding.message}：{finding.actual}",
                ]
            )
        self._render_table_grid(["类型", "文件", "工作表", "位置", "内容"], rows)

    @staticmethod
    def _format_table_sheet_detail(sheet) -> str:
        parts = [
            f"文件：{sheet.workbook}",
            f"工作表：{sheet.sheet}",
            f"数据范围：{sheet.max_row} 行 × {sheet.max_column} 列",
            f"识别表头行：第 {sheet.header_row} 行",
            "",
            "列标题：",
        ]
        parts.extend(f"{index}. {header}" for index, header in enumerate(sheet.headers, start=1))
        parts.extend(["", "合并单元格："])
        if sheet.merged_ranges:
            parts.extend(sheet.merged_ranges)
        else:
            parts.append("无")
        return "\n".join(parts)

    @staticmethod
    def _table_finding_location(finding) -> str:
        parts = []
        if finding.row is not None:
            parts.append(f"第 {finding.row} 行")
        if finding.column is not None:
            parts.append(f"第 {finding.column} 列")
        return "，".join(parts)

    @staticmethod
    def _format_table_merge_summary(report, output_path: Path) -> str:
        parts = [
            "汇总导出完成",
            f"输出文件：{output_path}",
            "",
            "统计：",
            f"副表数量：{report.stats.get('sources', 0)}",
            f"更新单元格：{report.stats.get('updated_cells', 0)}",
            f"追加内容条数：{report.stats.get('appended_values', 0)}",
            "",
            "提示：",
        ]
        if report.findings:
            for finding in report.findings[:20]:
                location = finding.sheet or Path(finding.workbook).name
                row = "" if finding.row is None else f" 第 {finding.row} 行"
                parts.append(f"[{finding.severity}] {location}{row}：{finding.message}")
            if len(report.findings) > 20:
                parts.append(f"还有 {len(report.findings) - 20} 条提示，请在上方结果列表中查看。")
        else:
            parts.append("无")
        return "\n".join(parts)

    def _clear_table_results(self) -> None:
        if self.table_preview_frame is not None:
            for child in self.table_preview_frame.winfo_children():
                child.destroy()
        self.table_details.clear()
        self._refresh_table_tree_placeholder()

    def _refresh_table_tree_placeholder(self) -> None:
        if self.table_tree_placeholder is None:
            return
        has_content = self.table_preview_frame is not None and bool(self.table_preview_frame.winfo_children())
        if has_content:
            self.table_tree_placeholder.place_forget()
        else:
            self.table_tree_placeholder.place(relx=0.5, rely=0.5, anchor="center")

    def _preview_table_path(self, path: Path, sheet_name: str | None = None) -> None:
        try:
            if path.resolve() not in self.table_workbook_infos:
                self._load_table_workbook_info(path)
            infos = self.table_workbook_infos.get(path.resolve()) or []
            sheet = sheet_name or (infos[0].sheet if infos else _first_sheet_name(path))
            header_row = next((info.header_row for info in infos if info.sheet == sheet), None)
            columns, rows = read_sheet_preview(path, sheet)
            merged_ranges = self._preview_merged_ranges(path, sheet)
            self._render_table_grid(
                ["行号", *columns],
                [[str(index), *row_values] for index, row_values in enumerate(rows, start=1)],
                header_row_index=header_row,
                merged_ranges=merged_ranges,
            )
            self.status_text.set(f"正在预览：{path.name} / {sheet}")
        except Exception as exc:
            self.status_text.set("表格预览失败")
            messagebox.showerror("表格预览失败", str(exc))

    def _preview_merged_ranges(self, path: Path, sheet_name: str) -> list[tuple[int, int, int, int]]:
        workbook = _load_workbook(path, data_only=False)
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        ranges: list[tuple[int, int, int, int]] = []
        for merged in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            if min_row > 80 or min_col > 50:
                continue
            ranges.append((min_row, min_col + 1, min(max_row, 80), min(max_col + 1, 51)))
        return ranges

    def _render_table_grid(
        self,
        headers: list[str],
        rows: list[list[str]],
        header_row_index: int | None = None,
        merged_ranges: list[tuple[int, int, int, int]] | None = None,
    ) -> None:
        self._clear_table_results()
        if self.table_preview_frame is None:
            return
        visible_headers = headers[:51]
        for column, header in enumerate(visible_headers):
            self._grid_cell(self.table_preview_frame, 0, column, header, header=True, width=self._preview_column_width(column, header))
        merged_lookup: dict[tuple[int, int], tuple[int, int]] = {}
        covered_cells: set[tuple[int, int]] = set()
        for min_row, min_col, max_row, max_col in merged_ranges or []:
            if max_row <= min_row and max_col <= min_col:
                continue
            merged_lookup[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    if (row, column) != (min_row, min_col):
                        covered_cells.add((row, column))
        for row_index, row_values in enumerate(rows[:80], start=1):
            is_header_row = header_row_index is not None and row_index == header_row_index
            values = row_values[: len(visible_headers)]
            for column in range(len(visible_headers)):
                if (row_index, column) in covered_cells:
                    continue
                value = values[column] if column < len(values) else ""
                rowspan, columnspan = merged_lookup.get((row_index, column), (1, 1))
                self._grid_cell(
                    self.table_preview_frame,
                    row_index,
                    column,
                    value,
                    header=is_header_row or column == 0,
                    width=self._preview_column_width(column, value),
                    rowspan=rowspan,
                    columnspan=columnspan,
                )
        self._refresh_table_tree_placeholder()

    @staticmethod
    def _preview_column_width(column: int, text: str) -> int:
        if column == 0:
            return 7
        return max(12, min(42, len(str(text)) + 4))

    @staticmethod
    def _grid_cell(
        parent: tk.Widget,
        row: int,
        column: int,
        text: str,
        *,
        header: bool = False,
        width: int = 12,
        rowspan: int = 1,
        columnspan: int = 1,
    ) -> None:
        bg = COLOR["surface_alt"] if header else COLOR["white"]
        fg = COLOR["accent"] if header else COLOR["text"]
        label = tk.Label(
            parent,
            text=str(text),
            bg=bg,
            fg=fg,
            width=width,
            anchor="nw",
            padx=6,
            pady=4,
            relief="solid",
            bd=1,
            justify="left",
            wraplength=max(90, min(360, width * 9)),
            font=(FONT_FAMILY, 9, "bold" if header else "normal"),
        )
        label.grid(row=row, column=column, rowspan=rowspan, columnspan=columnspan, sticky="nsew")


def main() -> int:
    _enable_dpi_awareness()
    root = tk.Tk()
    OfficeToolGUI(root)
    root.mainloop()
    return 0
