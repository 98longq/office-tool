"""Spreadsheet inspection helpers."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Iterable

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from .table_models import SheetInfo, TableReport


SUPPORTED_TABLE_INPUTS = {".xlsx"}


def _load_workbook(path: str | Path, *, data_only: bool = False):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("表格功能需要安装 openpyxl。") from exc
    return load_workbook(path, data_only=data_only)


def is_supported_table(path: Path) -> bool:
    return path.is_file() and not path.name.startswith("~$") and path.suffix.lower() in SUPPORTED_TABLE_INPUTS


def collect_table_inputs(paths: Iterable[str | Path], recursive: bool = True) -> list[Path]:
    inputs: list[Path] = []
    for raw in paths:
        path = Path(raw).expanduser().resolve()
        if path.is_file():
            if not is_supported_table(path):
                raise ValueError(f"不支持的表格格式: {path}")
            inputs.append(path)
            continue
        if path.is_dir():
            iterator = path.rglob("*") if recursive else path.glob("*")
            inputs.extend(sorted(file for file in iterator if is_supported_table(file)))
            continue
        raise FileNotFoundError(f"输入路径不存在: {path}")
    if not inputs:
        raise FileNotFoundError("未找到可处理的表格。当前支持 .xlsx")
    return inputs


def inspect_workbook(path: str | Path) -> list[SheetInfo]:
    source = Path(path).expanduser().resolve()
    workbook = _load_workbook(source, data_only=False)
    inspector = TableWorkbookInspector()
    return [inspector.inspect_sheet(source, sheet) for sheet in workbook.worksheets]


def read_sheet_preview(path: str | Path, sheet_name: str, max_rows: int = 80) -> tuple[list[str], list[list[str]]]:
    workbook = _load_workbook(path, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"工作表不存在: {sheet_name}")
    sheet = workbook[sheet_name]
    max_column = min(sheet.max_column, 50)
    columns = [_column_label(index) for index in range(1, max_column + 1)]
    rows: list[list[str]] = []
    for row_index in range(1, min(sheet.max_row, max_rows) + 1):
        rows.append([_clean_cell_value(sheet.cell(row_index, column).value) for column in range(1, max_column + 1)])
    return columns, rows


class TableWorkbookInspector:
    """Inspect xlsx workbooks without changing files."""

    def inspect_path(self, path: str | Path) -> TableReport:
        source = Path(path).expanduser().resolve()
        workbook = _load_workbook(source, data_only=False)
        report = TableReport()
        report.stats["workbooks"] = 1
        report.stats["workbook"] = str(source)
        for sheet in workbook.worksheets:
            report.sheets.append(self.inspect_sheet(source, sheet))
        report.stats["sheets"] = len(report.sheets)
        return report

    def inspect_many(self, paths: Iterable[str | Path]) -> TableReport:
        report = TableReport()
        for path in collect_table_inputs(paths):
            workbook_report = self.inspect_path(path)
            report.sheets.extend(workbook_report.sheets)
        report.stats["workbooks"] = len({sheet.workbook for sheet in report.sheets})
        report.stats["sheets"] = len(report.sheets)
        return report

    def inspect_sheet(self, workbook_path: str | Path, sheet: Worksheet) -> SheetInfo:
        header_row = detect_header_row(sheet)
        return SheetInfo(
            workbook=str(workbook_path),
            sheet=sheet.title,
            max_row=sheet.max_row,
            max_column=sheet.max_column,
            header_row=header_row,
            headers=headers_for_row(sheet, header_row),
            merged_ranges=[str(item) for item in sheet.merged_cells.ranges],
        )


def detect_header_row(sheet: Worksheet, scan_rows: int = 10) -> int:
    """Find the most likely header row near the top of a worksheet."""
    best_row = 1
    best_score = -1
    max_scan = min(max(1, scan_rows), sheet.max_row)
    for row_index in range(1, max_scan + 1):
        values = [_clean_cell_value(sheet.cell(row_index, column).value) for column in range(1, sheet.max_column + 1)]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue
        merged_title_penalty = 0
        if len(non_empty) == 1 and _row_has_wide_merge(sheet, row_index):
            merged_title_penalty = 4
        next_non_empty = 0
        if row_index < sheet.max_row:
            next_non_empty = sum(
                1
                for column in range(1, sheet.max_column + 1)
                if _clean_cell_value(sheet.cell(row_index + 1, column).value)
            )
        score = len(non_empty) * 3 + min(next_non_empty, len(non_empty)) - merged_title_penalty
        if score > best_score:
            best_score = score
            best_row = row_index
    return best_row


def headers_for_row(sheet: Worksheet, row_index: int) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        raw = _clean_cell_value(sheet.cell(row_index, column).value)
        header = raw or f"未命名列{column}"
        count = used.get(header, 0) + 1
        used[header] = count
        if count > 1:
            header = f"{header}_{count}"
        headers.append(header)
    return headers


def _clean_cell_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _column_label(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _row_has_wide_merge(sheet: Worksheet, row_index: int) -> bool:
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= row_index <= merged_range.max_row:
            if merged_range.max_col - merged_range.min_col >= 1:
                return True
    return False
