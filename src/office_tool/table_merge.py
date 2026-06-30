"""Column-based spreadsheet merge engine."""

from __future__ import annotations

import re
import unicodedata
from copy import copy
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import TYPE_CHECKING

from .table_audit import _load_workbook, detect_header_row, headers_for_row
from .table_models import SourceColumnMapping, TableMergeOptions, TableReport

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class SourceValue:
    value: str
    workbook: Path
    sheet: str
    row: int


def merge_by_columns(options: TableMergeOptions) -> TableReport:
    """Merge source column values into a master sheet by matching key columns."""
    master_path = Path(options.master_path).expanduser().resolve()
    output_path = Path(options.output_path).expanduser().resolve()
    master_wb = _load_workbook(master_path, data_only=False)
    if options.master_sheet not in master_wb.sheetnames:
        raise KeyError(f"主表不存在工作表: {options.master_sheet}")
    master_ws = master_wb[options.master_sheet]
    master_header_row = options.master_header_row or detect_header_row(master_ws)
    master_headers = headers_for_row(master_ws, master_header_row)
    master_key_col = resolve_column(master_ws, master_header_row, options.master_key_column)
    master_target_col = resolve_column(master_ws, master_header_row, options.master_target_column)

    report = TableReport()
    report.stats["master"] = str(master_path)
    report.stats["output"] = str(output_path)
    report.stats["sources"] = len(options.sources)
    report.stats["updated_cells"] = 0
    report.stats["appended_values"] = 0
    report.sheets.append(
        _sheet_info(str(master_path), master_ws, master_header_row, master_headers)
    )

    source_values = _collect_source_values(options, report)
    match_plan = _build_key_match_plan(master_ws, master_header_row, master_key_col, source_values, options, report)
    if report.count("error"):
        raise ValueError("模糊匹配存在一对多或多对一歧义，请调整匹配度或关闭模糊匹配后重试。")
    used_source_keys: set[tuple[str, str, int]] = set()
    for row in range(master_header_row + 1, master_ws.max_row + 1):
        key = normalized_cell_text(master_ws.cell(row, master_key_col).value, options.normalize_keys)
        if not key:
            report.add_finding(
                "empty_master_key",
                "warning",
                "主表校验列为空，已跳过该行。",
                workbook=master_path,
                sheet=master_ws.title,
                row=row,
                column=master_key_col,
            )
            continue
        matches = match_plan.get(row, [])
        if not matches:
            report.add_finding(
                "unmatched_master_key",
                "info",
                "主表校验值未在副表中找到匹配数据。",
                workbook=master_path,
                sheet=master_ws.title,
                row=row,
                column=master_key_col,
                actual=master_ws.cell(row, master_key_col).value or "",
            )
            continue
        target_cell = master_ws.cell(row, master_target_col)
        existing = split_cell_values(target_cell.value, options.separator)
        additions: list[str] = []
        for match in matches:
            used_source_keys.add((str(match.workbook), match.sheet, match.row))
            if match.value and match.value not in existing and match.value not in additions:
                additions.append(match.value)
        if not additions:
            continue
        target_cell.value = options.separator.join(existing + additions) if existing else options.separator.join(additions)
        report.stats["updated_cells"] += 1
        report.stats["appended_values"] += len(additions)
        if len(additions) > 1 or existing:
            report.add_finding(
                "merged_multiple_values",
                "info",
                "同一主表单元格合并了多条不同内容。",
                workbook=master_path,
                sheet=master_ws.title,
                row=row,
                column=master_target_col,
                actual=target_cell.value or "",
            )

    for key, values in source_values.items():
        for value in values:
            marker = (str(value.workbook), value.sheet, value.row)
            if marker not in used_source_keys:
                report.add_finding(
                    "unused_source_value",
                    "info",
                    "副表数据未匹配到主表行。",
                    workbook=value.workbook,
                    sheet=value.sheet,
                    row=value.row,
                    actual=f"{key}: {value.value}",
                )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    master_wb.save(output_path)
    return report


def _build_key_match_plan(
    master_ws: "Worksheet",
    master_header_row: int,
    master_key_col: int,
    source_values: dict[str, list[SourceValue]],
    options: TableMergeOptions,
    report: TableReport,
) -> dict[int, list[SourceValue]]:
    if not options.fuzzy_match:
        plan: dict[int, list[SourceValue]] = {}
        for row in range(master_header_row + 1, master_ws.max_row + 1):
            key = normalized_cell_text(master_ws.cell(row, master_key_col).value, options.normalize_keys)
            if key:
                plan[row] = source_values.get(key, [])
        return plan

    threshold = max(0, min(100, int(options.fuzzy_threshold)))
    master_keys: dict[int, str] = {}
    best_by_row: dict[int, tuple[str, float]] = {}
    rows_by_source_key: dict[str, list[int]] = {}
    for row in range(master_header_row + 1, master_ws.max_row + 1):
        key = normalized_cell_text(master_ws.cell(row, master_key_col).value, options.normalize_keys)
        if not key:
            continue
        master_keys[row] = key
        candidates: list[tuple[str, float]] = []
        for source_key in source_values:
            score = _similarity_score(key, source_key)
            if score >= threshold:
                candidates.append((source_key, score))
        if not candidates:
            continue
        candidates.sort(key=lambda item: item[1], reverse=True)
        top_score = candidates[0][1]
        top_keys = [item[0] for item in candidates if item[1] == top_score]
        if len(top_keys) > 1:
            report.add_finding(
                "ambiguous_fuzzy_match",
                "error",
                "模糊匹配出现一对多候选，已停止汇总。",
                sheet=master_ws.title,
                row=row,
                column=master_key_col,
                actual=key,
                suggestion="提高匹配度，或手动统一主表和副表校验文字。",
            )
            continue
        source_key = top_keys[0]
        if len(source_values.get(source_key, [])) > 1:
            report.add_finding(
                "ambiguous_fuzzy_match",
                "error",
                "模糊匹配命中多条副表数据，已停止汇总。",
                sheet=master_ws.title,
                row=row,
                column=master_key_col,
                actual=f"{key} ≈ {source_key}",
                suggestion="提高匹配度，或先处理副表重复校验值。",
            )
            continue
        best_by_row[row] = (source_key, top_score)
        rows_by_source_key.setdefault(source_key, []).append(row)

    for source_key, rows in rows_by_source_key.items():
        if len(rows) <= 1:
            continue
        report.add_finding(
            "ambiguous_fuzzy_match",
            "error",
            "模糊匹配出现多对一候选，已停止汇总。",
            sheet=master_ws.title,
            actual=source_key,
            suggestion="提高匹配度，或手动统一主表和副表校验文字。",
        )

    if report.count("error"):
        return {}

    plan: dict[int, list[SourceValue]] = {}
    for row, (source_key, score) in best_by_row.items():
        plan[row] = source_values.get(source_key, [])
        if master_keys[row] != source_key:
            report.add_finding(
                "fuzzy_key_match",
                "info",
                "已使用模糊匹配关联主表和副表校验值。",
                sheet=master_ws.title,
                row=row,
                column=master_key_col,
                actual=f"{master_keys[row]} ≈ {source_key} ({score:.0f}%)",
            )
    return plan


def _similarity_score(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    ratio = SequenceMatcher(None, left, right).ratio()
    if left in right or right in left:
        ratio = max(ratio, min(len(left), len(right)) / max(len(left), len(right)))
    return ratio * 100


def merge_same_layout(
    master_path: str | Path,
    source_paths: list[str | Path],
    output_path: str | Path,
    *,
    master_sheet: str | None = None,
    source_sheet: str | None = None,
    separator: str = "\n",
) -> TableReport:
    """Merge same-layout workbooks by cell position, preserving the master workbook."""
    master_path = Path(master_path).expanduser().resolve()
    output_path = Path(output_path).expanduser().resolve()
    master_wb = _load_workbook(master_path, data_only=False)
    master_ws = master_wb[master_sheet] if master_sheet else master_wb.worksheets[0]

    report = TableReport()
    report.stats["master"] = str(master_path)
    report.stats["output"] = str(output_path)
    report.stats["sources"] = len(source_paths)
    report.stats["updated_cells"] = 0
    report.stats["appended_values"] = 0
    report.stats["appended_rows"] = 0
    report.stats["skipped_formula_cells"] = 0
    report.stats["layout_warnings"] = 0
    header_row = detect_header_row(master_ws)
    master_headers = headers_for_row(master_ws, header_row)
    master_merged_ranges = _merged_range_texts(master_ws)
    template_max_row = master_ws.max_row
    template_max_col = master_ws.max_column
    append_rows_mode = not _sheet_has_data_below_header(master_ws, header_row)
    report.stats["merge_mode"] = "append_rows" if append_rows_mode else "cell_position"
    report.sheets.append(_sheet_info(str(master_path), master_ws, header_row, master_headers))

    for raw_source in source_paths:
        source_path = Path(raw_source).expanduser().resolve()
        workbook = _load_workbook(source_path, data_only=False)
        if source_sheet:
            if source_sheet not in workbook.sheetnames:
                report.add_finding(
                    "missing_source_sheet",
                    "warning",
                    "副表不存在指定工作表，已使用第一个工作表。",
                    workbook=source_path,
                    sheet=source_sheet,
                    suggestion="如该副表格式特殊，请在高级设置中单独配置。",
                )
                sheet = workbook.worksheets[0]
            else:
                sheet = workbook[source_sheet]
        else:
            sheet = workbook.worksheets[0]
        source_header_row = detect_header_row(sheet)
        source_headers = headers_for_row(sheet, source_header_row)
        report.sheets.append(_sheet_info(str(source_path), sheet, source_header_row, source_headers))
        _warn_same_layout_mismatches(
            report,
            master_ws=master_ws,
            source_ws=sheet,
            source_path=source_path,
            master_header_row=header_row,
            source_header_row=source_header_row,
            master_headers=master_headers,
            source_headers=source_headers,
            master_merged_ranges=master_merged_ranges,
            compare_size=not append_rows_mode,
        )
        if append_rows_mode:
            source_values_seen = 0
            for row in range(source_header_row + 1, sheet.max_row + 1):
                if _append_source_row(master_ws, sheet, row, report, max_column=template_max_col):
                    source_values_seen += 1
                    report.stats["appended_rows"] += 1
            if source_values_seen == 0:
                report.add_finding(
                    "empty_source_sheet",
                    "info",
                    "副表工作表没有可汇总的明细行，已跳过。",
                    workbook=source_path,
                    sheet=sheet.title,
                )
            continue
        max_row = min(template_max_row, sheet.max_row)
        max_col = min(template_max_col, sheet.max_column)
        source_values_seen = 0
        for row in range(1, max_row + 1):
            for column in range(1, max_col + 1):
                source_cell = sheet.cell(row, column)
                value = normalized_cell_text(source_cell.value, normalize=False)
                if not value:
                    continue
                source_values_seen += 1
                if row <= header_row:
                    continue
                if source_cell.data_type == "f":
                    report.stats["skipped_formula_cells"] += 1
                    continue
                target_cell = master_ws.cell(row, column)
                if target_cell.data_type == "f":
                    report.stats["skipped_formula_cells"] += 1
                    continue
                existing = split_cell_values(target_cell.value, separator)
                if value in existing:
                    continue
                if existing and not _looks_like_aggregate_cell(target_cell.value, separator) and not _is_fillable_same_layout_column(master_headers, column):
                    report.add_finding(
                        "protected_template_cell",
                        "warning",
                        "主表该单元格已有模板内容，已跳过副表差异值。",
                        workbook=master_path,
                        sheet=master_ws.title,
                        row=row,
                        column=column,
                        actual=target_cell.value or "",
                        suggestion="如该列确实需要汇总，请使用按列汇总并指定填入列。",
                    )
                    continue
                target_cell.value = separator.join(existing + [value]) if existing else value
                report.stats["updated_cells"] += 1
                report.stats["appended_values"] += 1
        for row in range(template_max_row + 1, sheet.max_row + 1):
            if _append_source_row(master_ws, sheet, row, report):
                source_values_seen += 1
                report.stats["appended_rows"] += 1
        if source_values_seen == 0:
            report.add_finding(
                "empty_source_sheet",
                "info",
                "副表工作表为空，已跳过。",
                workbook=source_path,
                sheet=sheet.title,
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    master_wb.save(output_path)
    return report


def _append_source_row(master_ws: "Worksheet", source_ws: "Worksheet", source_row: int, report: TableReport, *, max_column: int | None = None) -> bool:
    last_column = max_column or source_ws.max_column
    values = [source_ws.cell(source_row, column).value for column in range(1, last_column + 1)]
    if not any(normalized_cell_text(value, normalize=False) for value in values):
        return False
    target_row = master_ws.max_row + 1
    for column in range(1, last_column + 1):
        source_cell = source_ws.cell(source_row, column)
        if source_cell.data_type == "f":
            report.stats["skipped_formula_cells"] += 1
            continue
        target_cell = master_ws.cell(target_row, column)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
    source_height = source_ws.row_dimensions[source_row].height
    if source_height:
        master_ws.row_dimensions[target_row].height = source_height
    return True


def _warn_same_layout_mismatches(
    report: TableReport,
    *,
    master_ws: "Worksheet",
    source_ws: "Worksheet",
    source_path: Path,
    master_header_row: int,
    source_header_row: int,
    master_headers: list[str],
    source_headers: list[str],
    master_merged_ranges: set[str],
    compare_size: bool = True,
) -> None:
    if master_header_row != source_header_row:
        _add_layout_warning(
            report,
            source_path,
            source_ws.title,
            f"副表表头行与主表不一致：主表第 {master_header_row} 行，副表第 {source_header_row} 行。",
        )
    if _compact_headers(master_headers) != _compact_headers(source_headers):
        _add_layout_warning(report, source_path, source_ws.title, "副表表头与主表不完全一致，已按交叉区域谨慎汇总。")
    if _merged_range_texts(source_ws) != master_merged_ranges:
        _add_layout_warning(report, source_path, source_ws.title, "副表合并单元格结构与主表不完全一致。")
    if compare_size and (source_ws.max_row != master_ws.max_row or source_ws.max_column != master_ws.max_column):
        _add_layout_warning(
            report,
            source_path,
            source_ws.title,
            f"副表尺寸与主表不一致：主表 {master_ws.max_row} 行×{master_ws.max_column} 列，副表 {source_ws.max_row} 行×{source_ws.max_column} 列。",
        )


def _add_layout_warning(report: TableReport, workbook: Path, sheet: str, message: str) -> None:
    report.stats["layout_warnings"] = int(report.stats.get("layout_warnings", 0)) + 1
    report.add_finding(
        "layout_mismatch",
        "warning",
        message,
        workbook=workbook,
        sheet=sheet,
        suggestion="建议先预览结果；如果模板差异较大，请改用按列汇总。",
    )


def _compact_headers(headers: list[str]) -> list[str]:
    return [normalized_cell_text(header) for header in headers if normalized_cell_text(header)]


def _merged_range_texts(sheet: "Worksheet") -> set[str]:
    return {str(item) for item in sheet.merged_cells.ranges}


def _sheet_has_data_below_header(sheet: "Worksheet", header_row: int) -> bool:
    for row in range(header_row + 1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            if normalized_cell_text(sheet.cell(row, column).value, normalize=False):
                return True
    return False


def _looks_like_aggregate_cell(value, separator: str) -> bool:
    text = "" if value is None else str(value)
    return separator in text


def _is_fillable_same_layout_column(headers: list[str], column: int) -> bool:
    if column < 1 or column > len(headers):
        return False
    header = normalized_cell_text(headers[column - 1])
    if not header:
        return False
    protected_words = ("序号", "编号", "任务", "事项", "名称", "标题", "部门", "单位")
    fillable_words = ("回复", "反馈", "落实", "办理", "情况", "进展", "结果", "备注", "意见", "说明", "内容")
    if any(word in header for word in fillable_words):
        return True
    if any(word in header for word in protected_words):
        return False
    return False


def resolve_column(sheet: "Worksheet", header_row: int, column: str | int) -> int:
    if isinstance(column, int):
        if column < 1:
            raise ValueError("列序号必须从 1 开始。")
        return column
    text = str(column).strip()
    if not text:
        raise ValueError("列不能为空。")
    if re.fullmatch(r"[A-Za-z]{1,3}", text):
        try:
            from openpyxl.utils.cell import column_index_from_string
        except ImportError as exc:
            raise RuntimeError("表格功能需要安装 openpyxl。") from exc
        return column_index_from_string(text.upper())
    headers = headers_for_row(sheet, header_row)
    for index, header in enumerate(headers, start=1):
        if header == text:
            return index
    raise KeyError(f"未找到列标题: {text}")


def normalized_cell_text(value, normalize: bool = True) -> str:
    if value is None:
        return ""
    text = str(value)
    if not normalize:
        return text.strip()
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def split_cell_values(value, separator: str) -> list[str]:
    text = "" if value is None else str(value)
    if not text.strip():
        return []
    return [part.strip() for part in text.split(separator) if part.strip()]


def _collect_source_values(options: TableMergeOptions, report: TableReport) -> dict[str, list[SourceValue]]:
    values_by_key: dict[str, list[SourceValue]] = {}
    for source in options.sources:
        source_path = Path(source.path).expanduser().resolve()
        workbook = _load_workbook(source_path, data_only=False)
        if source.sheet not in workbook.sheetnames:
            raise KeyError(f"副表不存在工作表: {source.sheet}")
        sheet = workbook[source.sheet]
        header_row = source.header_row or detect_header_row(sheet)
        report.sheets.append(_sheet_info(str(source_path), sheet, header_row, headers_for_row(sheet, header_row)))
        key_col = resolve_column(sheet, header_row, source.key_column)
        value_col = resolve_column(sheet, header_row, source.value_column)
        seen_keys: set[str] = set()
        for row in range(header_row + 1, sheet.max_row + 1):
            key = normalized_cell_text(sheet.cell(row, key_col).value, options.normalize_keys)
            value = normalized_cell_text(sheet.cell(row, value_col).value, normalize=False)
            if not key:
                continue
            if key in seen_keys:
                report.add_finding(
                    "duplicate_source_key",
                    "warning",
                    "同一副表内校验值重复，后续会按多条候选合并。",
                    workbook=source_path,
                    sheet=sheet.title,
                    row=row,
                    column=key_col,
                    actual=key,
                )
            seen_keys.add(key)
            if not value:
                continue
            bucket = values_by_key.setdefault(key, [])
            if value not in {item.value for item in bucket}:
                bucket.append(SourceValue(value=value, workbook=source_path, sheet=sheet.title, row=row))
    return values_by_key


def _sheet_info(workbook_path: str, sheet: "Worksheet", header_row: int, headers: list[str]):
    from .table_models import SheetInfo

    return SheetInfo(
        workbook=workbook_path,
        sheet=sheet.title,
        max_row=sheet.max_row,
        max_column=sheet.max_column,
        header_row=header_row,
        headers=headers,
        merged_ranges=[str(item) for item in sheet.merged_cells.ranges],
    )
